#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
TEST DEFINITIVO - SISTEMA DE ANÁLISIS FINANCIERO SMV
====================================================

Test final consolidado que prueba todo el sistema de análisis financiero:
✅ Simula el flujo completo de Django
✅ Valida todas las funciones de análisis  
✅ Genera reportes automáticamente
✅ Abre resultados para inspección visual
✅ Incluye métricas de rendimiento

Uso: python test_definitivo.py
"""

import os
import sys
import shutil
import traceback
from pathlib import Path
from datetime import datetime

# Configuración del proyecto
PROJECT_ROOT = Path(__file__).parent
sys.path.insert(0, str(PROJECT_ROOT))
sys.path.insert(0, str(PROJECT_ROOT / "SMV_APP"))

# Importar funciones de análisis
try:
    from SMV_APP.analisis import (
        formato_xls_xlsx, union_archivos, analisis_VH, analisis_Ratios,
        graficosRatios, analisisVertical, analisisHorizontal, analisisRatiosCalculo,
        renombrar, crear_graficos_analisis
    )
    from openpyxl import load_workbook
    print("✅ Todos los módulos importados correctamente")
except ImportError as e:
    print(f"❌ ERROR CRÍTICO - Falta módulo: {e}")
    print("💡 Ejecuta: pip install openpyxl pandas")
    sys.exit(1)

class TestDefinitivoAnalisisFinanciero:
    """Test definitivo para el sistema de análisis financiero"""
    
    def __init__(self):
        self.project_root = PROJECT_ROOT
        self.descargas_path = self.project_root / "descargas_smv"
        self.resultados_path = self.project_root / "RESULTADOS_TEST_FINAL"
        self.start_time = datetime.now()
        self.metricas = {
            'archivos_procesados': 0,
            'funciones_exitosas': 0,
            'funciones_fallidas': 0,
            'tiempo_total': 0,
            'archivo_final': None
        }
        
    def limpiar_resultados_anteriores(self):
        """Limpiar resultados de pruebas anteriores"""
        if self.resultados_path.exists():
            shutil.rmtree(self.resultados_path)
        self.resultados_path.mkdir(exist_ok=True)
        print("🧹 Resultados anteriores limpiados")
        
    def verificar_prerequisitos(self):
        """Verificar que todo esté listo para las pruebas"""
        print("\n🔍 VERIFICANDO PRERREQUISITOS...")
        
        # Verificar directorio de datos
        if not self.descargas_path.exists():
            print(f"❌ No existe: {self.descargas_path}")
            return False
            
        # Buscar empresas disponibles
        empresas = [d for d in self.descargas_path.iterdir() if d.is_dir()]
        if not empresas:
            print("❌ No hay empresas para probar")
            return False
            
        # Verificar archivos Excel
        total_archivos = 0
        for empresa in empresas:
            archivos = list(empresa.glob("*.xlsx")) + list(empresa.glob("*.xls"))
            total_archivos += len(archivos)
            
        if total_archivos == 0:
            print("❌ No hay archivos Excel para procesar")
            return False
            
        print(f"✅ {len(empresas)} empresas encontradas")
        print(f"✅ {total_archivos} archivos Excel disponibles")
        return True
        
    def seleccionar_empresa_optima(self):
        """Seleccionar la mejor empresa para pruebas"""
        empresas = [d for d in self.descargas_path.iterdir() if d.is_dir()]
        
        mejor_empresa = None
        max_puntos = 0
        
        for empresa in empresas:
            archivos_xlsx = list(empresa.glob("*.xlsx"))
            archivos_xls = list(empresa.glob("*.xls"))
            
            # Puntuación basada en variedad y cantidad de archivos
            puntos = len(archivos_xlsx) * 3 + len(archivos_xls) * 2
            
            # Bonus si tiene archivos de múltiples años
            años = set()
            for archivo in archivos_xlsx + archivos_xls:
                for parte in archivo.name.split('-'):
                    if parte.isdigit() and len(parte) == 4:
                        años.add(int(parte))
            puntos += len(años) * 5
            
            if puntos > max_puntos:
                max_puntos = puntos
                mejor_empresa = empresa
                
        return mejor_empresa
        
    def preparar_archivos_test(self, empresa_path):
        """Preparar archivos para el test"""
        print(f"\n📋 PREPARANDO ARCHIVOS DE {empresa_path.name}...")
        
        # Crear directorio para esta empresa
        empresa_test_path = self.resultados_path / empresa_path.name
        empresa_test_path.mkdir(exist_ok=True)
        
        # Copiar archivos Excel
        archivos_copiados = []
        for extension in ['*.xlsx', '*.xls']:
            for archivo in empresa_path.glob(extension):
                destino = empresa_test_path / archivo.name
                shutil.copy2(archivo, destino)
                archivos_copiados.append(destino)
                print(f"   📄 {archivo.name}")
                
        self.metricas['archivos_procesados'] = len(archivos_copiados)
        return empresa_test_path, archivos_copiados
        
    def simular_flujo_django(self, empresa_test_path, archivos):
        """Simular exactamente el flujo que hace Django"""
        print(f"\n🔄 SIMULANDO FLUJO COMPLETO DE DJANGO...")
        
        # 1. Procesar archivos .xls a .xlsx si es necesario
        print("   🔧 Paso 1: Convertir XLS → XLSX...")
        formato_xls_xlsx(str(empresa_test_path))
        
        # 2. Organizar archivos por año
        print("   📅 Paso 2: Organizar por años...")
        archivos_por_anio = {}
        
        for archivo in empresa_test_path.glob("*.xlsx"):
            # Extraer año del nombre
            for parte in archivo.name.split('-'):
                if parte.isdigit() and len(parte) == 4:
                    year = int(parte)
                    archivos_por_anio[year] = archivo
                    break
                    
        if not archivos_por_anio:
            print("   ❌ No se pudieron extraer años")
            return None
            
        años_ordenados = sorted(archivos_por_anio.keys(), reverse=True)
        print(f"   ✅ Años encontrados: {años_ordenados}")
        
        # 3. Crear archivo base (más reciente)
        print("   📊 Paso 3: Crear archivo de análisis...")
        archivo_base = archivos_por_anio[años_ordenados[0]]
        
        carpeta_analisis = empresa_test_path / "ANALISIS"
        carpeta_analisis.mkdir(exist_ok=True)
        
        nombre_final = f"ANALISIS_DEFINITIVO_{empresa_test_path.name}.xlsx"
        archivo_analisis = carpeta_analisis / nombre_final
        shutil.copy2(archivo_base, archivo_analisis)
        
        # 4. Unir archivos adicionales
        print("   🔗 Paso 4: Unir datos históricos...")
        for i, año in enumerate(años_ordenados[1:], start=1):
            if i <= 4:  # Máximo 4 archivos adicionales
                try:
                    archivo_origen = archivos_por_anio[año]
                    columna = 4 + i
                    union_archivos(str(archivo_origen), str(archivo_analisis), columna)
                    print(f"      ✅ {año} → Columna {columna}")
                except Exception as e:
                    print(f"      ⚠️ Error uniendo {año}: {e}")
                    
        return archivo_analisis
        
    def ejecutar_analisis_completo(self, archivo_analisis):
        """Ejecutar todas las funciones de análisis en orden correcto"""
        print(f"\n📈 EJECUTANDO ANÁLISIS FINANCIERO COMPLETO...")
        
        funciones_analisis = [
            ("analisis_VH", analisis_VH, "Crear estructura de análisis V&H"),
            ("analisis_Ratios", analisis_Ratios, "Crear hoja de ratios"),
            ("graficosRatios", graficosRatios, "Preparar gráficos"),
            ("analisisVertical", analisisVertical, "Calcular análisis vertical"),
            ("analisisHorizontal", analisisHorizontal, "Calcular análisis horizontal"),
            ("analisisRatiosCalculo", analisisRatiosCalculo, "Calcular valores ratios"),
            ("renombrar", renombrar, "Renombrar hojas finales"),
            ("crear_graficos_analisis", crear_graficos_analisis, "Crear gráficos de análisis V&H")
        ]
        
        resultados_funciones = {}
        
        for nombre, funcion, descripcion in funciones_analisis:
            try:
                print(f"   🔬 {nombre}: {descripcion}...")
                funcion(str(archivo_analisis))
                print(f"   ✅ {nombre} completado")
                resultados_funciones[nombre] = True
                self.metricas['funciones_exitosas'] += 1
                
            except Exception as e:
                print(f"   ❌ {nombre} falló: {e}")
                resultados_funciones[nombre] = False
                self.metricas['funciones_fallidas'] += 1
                
        return resultados_funciones
        
    def validar_archivo_final(self, archivo_analisis):
        """Validar que el archivo final esté correcto"""
        print(f"\n🔍 VALIDANDO ARCHIVO FINAL...")
        
        try:
            wb = load_workbook(archivo_analisis, read_only=True)
            
            print(f"   📄 Archivo: {archivo_analisis.name}")
            print(f"   📏 Tamaño: {archivo_analisis.stat().st_size / 1024:.1f} KB")
            print(f"   📑 Hojas ({len(wb.sheetnames)}):")
            
            hojas_esperadas = [
                "Estado de Situación Financiera",
                "Estado de Resultados", 
                "Estado de Patrimonio Neto",
                "Estado de Flujo de Efectivo",
                "Ratios Financieros",
                "Gráficos de Ratios Financieros"
            ]
            
            hojas_encontradas = 0
            for i, hoja in enumerate(wb.sheetnames, 1):
                esta_esperada = hoja in hojas_esperadas
                status = "✅" if esta_esperada else "⚠️"
                print(f"      {i}. {status} {hoja}")
                if esta_esperada:
                    hojas_encontradas += 1
                    
            wb.close()
            
            porcentaje_completitud = (hojas_encontradas / len(hojas_esperadas)) * 100
            print(f"   📊 Completitud: {hojas_encontradas}/{len(hojas_esperadas)} ({porcentaje_completitud:.1f}%)")
            
            self.metricas['archivo_final'] = str(archivo_analisis)
            return porcentaje_completitud >= 80
            
        except Exception as e:
            print(f"   ❌ Error validando: {e}")
            return False
            
    def generar_reporte_final(self, resultados_funciones, validacion_exitosa):
        """Generar reporte completo de resultados"""
        end_time = datetime.now()
        self.metricas['tiempo_total'] = (end_time - self.start_time).total_seconds()
        
        print("\n" + "=" * 70)
        print("📊 REPORTE FINAL - TEST DEFINITIVO")
        print("=" * 70)
        
        print(f"⏱️  Tiempo total: {self.metricas['tiempo_total']:.2f} segundos")
        print(f"📄 Archivos procesados: {self.metricas['archivos_procesados']}")
        print(f"✅ Funciones exitosas: {self.metricas['funciones_exitosas']}")
        print(f"❌ Funciones fallidas: {self.metricas['funciones_fallidas']}")
        
        if self.metricas['archivo_final']:
            print(f"📁 Archivo final: {self.metricas['archivo_final']}")
            
        print("\n📈 Detalle por función:")
        for func, exitosa in resultados_funciones.items():
            status = "✅" if exitosa else "❌"
            print(f"   {status} {func}")
            
        # Calcular puntuación general
        total_funciones = len(resultados_funciones)
        exitosas = sum(1 for success in resultados_funciones.values() if success)
        puntuacion = (exitosas / total_funciones) * 100 if total_funciones > 0 else 0
        
        print(f"\n🎯 PUNTUACIÓN FINAL: {exitosas}/{total_funciones} ({puntuacion:.1f}%)")
        
        if puntuacion >= 90 and validacion_exitosa:
            print("🏆 ¡EXCELENTE! Sistema funcionando perfectamente")
            return "EXCELENTE"
        elif puntuacion >= 70:
            print("✅ BUENO: Sistema funcionando correctamente") 
            return "BUENO"
        else:
            print("⚠️ NECESITA ATENCIÓN: Revisar errores")
            return "ATENCION"
            
    def abrir_resultados(self):
        """Abrir automáticamente los resultados"""
        if self.metricas['archivo_final']:
            try:
                print(f"\n🚀 Abriendo archivo final...")
                if os.name == 'nt':  # Windows
                    os.startfile(self.metricas['archivo_final'])
                else:  # Linux/Mac
                    import subprocess
                    subprocess.call(['xdg-open', self.metricas['archivo_final']])
                print("✅ Archivo abierto para inspección")
            except Exception as e:
                print(f"⚠️ No se pudo abrir automáticamente: {e}")
                print(f"📁 Abre manualmente: {self.metricas['archivo_final']}")
                
    def ejecutar_test_completo(self):
        """Método principal que ejecuta todo el test"""
        print("🧪 TEST DEFINITIVO DEL SISTEMA DE ANÁLISIS FINANCIERO")
        print("=" * 60)
        
        # Paso 1: Preparación
        self.limpiar_resultados_anteriores()
        if not self.verificar_prerequisitos():
            return False
            
        # Paso 2: Selección de empresa
        empresa_path = self.seleccionar_empresa_optima()
        if not empresa_path:
            print("❌ No se pudo seleccionar empresa")
            return False
            
        print(f"🏢 Empresa seleccionada: {empresa_path.name}")
        
        # Paso 3: Preparar archivos
        empresa_test_path, archivos = self.preparar_archivos_test(empresa_path)
        
        # Paso 4: Simular flujo Django
        archivo_analisis = self.simular_flujo_django(empresa_test_path, archivos)
        if not archivo_analisis:
            print("❌ Falló la simulación del flujo Django")
            return False
            
        # Paso 5: Ejecutar análisis
        resultados_funciones = self.ejecutar_analisis_completo(archivo_analisis)
        
        # Paso 6: Validar resultado
        validacion_exitosa = self.validar_archivo_final(archivo_analisis)
        
        # Paso 7: Generar reporte
        resultado_final = self.generar_reporte_final(resultados_funciones, validacion_exitosa)
        
        # Paso 8: Abrir resultados
        self.abrir_resultados()
        
        return resultado_final in ["EXCELENTE", "BUENO"]

def main():
    """Función principal del test definitivo"""
    print("🔬 INICIANDO TEST DEFINITIVO...")
    print("💡 Este test prueba TODO el sistema de análisis financiero")
    print()
    
    try:
        tester = TestDefinitivoAnalisisFinanciero()
        exito = tester.ejecutar_test_completo()
        
        if exito:
            print("\n🎉 ¡TEST DEFINITIVO COMPLETADO EXITOSAMENTE!")
            print("📁 Revisa la carpeta RESULTADOS_TEST_FINAL/")
        else:
            print("\n⚠️ Test completado con algunas fallas")
            print("🔧 Revisa los errores reportados arriba")
            
    except KeyboardInterrupt:
        print("\n👋 Test cancelado por el usuario")
    except Exception as e:
        print(f"\n❌ ERROR INESPERADO: {e}")
        traceback.print_exc()

if __name__ == "__main__":
    main()