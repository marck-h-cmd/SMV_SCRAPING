import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries


# Color de relleno de encabezado (Azul Oscuro/Teal)
HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")

# Color de relleno de encabezado (Azul Oscuro/Teal)
ENCABEZADO_NARANJA = PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid")

ENCABEZADO_PURPURA = PatternFill(start_color="BE78F8", end_color="BE78F8", fill_type="solid")

ENCABEZADO_CELESTE = PatternFill(start_color="65FFD7", end_color="65FFD7", fill_type="solid")

# Color de relleno para la sección principal (Gris Claro)
SECTION_FILL = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

# Fuente para encabezados (Blanco y Negrita)
HEADER_FONT = Font(size=9, bold=True, color="FFFFFF")
Contenido = Font(name="Franklin Gothic Medium Cond", size=9)

# Estilo de borde (borde fino para todas las celdas de datos)
THIN_SIDE = Side(style='thin', color="000000")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

fuente_titulo = Font(size=13.5, bold=True)
negrita = Font(bold=True)
cuentas = Alignment(horizontal='center', vertical='center')

# Formato de SIMBOLO DECIMAL (.), MILES (,) SIMBOLO NEGATIVO (-)
FORMATO_NUMERICO_FINANCIERO = '#,##0.00;-#,##0.00'
FORMATO_NUMERICO = '#,##0;-#,##0.'
FORMATO_PORCENTAJE_DOS_DECIMALES = '0.00%'

# -----------------------------
def formato_xls_xlsx(CARPETA_FINANCIEROS):
    #FORMATO DE LOS ESTADOS FINANCIEROS - Ultimos 5 años
    for archivo in os.listdir(CARPETA_FINANCIEROS):
        if archivo.endswith(".xls"):
            path_xls = os.path.join(CARPETA_FINANCIEROS, archivo)
            path_xlsx = path_xls + "x"
            dir_path = os.path.dirname(path_xls)    #Ruta del excel
            nombre = os.path.basename(dir_path)    #Nombre de la empresa
            nombreEmpresa = nombre.replace('_', ' ')
            tablas = pd.read_html(path_xls)

            # 1. GUARDAR DATAFRAME
            with pd.ExcelWriter(path_xlsx, engine="openpyxl") as writer:
                for i, df in enumerate(tablas, start=1):

                    if df.shape[1] >= 4 and i!=3:   #TODOS MENOS LA DE PATRIMONIO O FIRMANTES ELIMINAN LA COLUMNA B
                        df_cleaned = df.drop(df.columns[[1, 3]], axis=1, errors='ignore')
                    else:
                        df_cleaned = df

                    def clean_and_coerce(series):
                        s = series.astype(str)
                        s = s.str.replace('(', '-', regex=False).str.replace(')', '', regex=False)
                        s = s.str.replace(',', '', regex=False)
                        return pd.to_numeric(s, errors='coerce')

                    if df_cleaned.shape[1] >= 3 and (i!=3 and i!=6):
                        df_cleaned.iloc[:, [1, 2]] = df_cleaned.iloc[:, [1, 2]].apply(clean_and_coerce)
                    else:
                        df_cleaned.iloc[:, 2:df_cleaned.shape[1]] = df_cleaned.iloc[:, 2:df_cleaned.shape[1]].apply(clean_and_coerce)

                    df_cleaned.to_excel(
                        writer, 
                        sheet_name=f"Hoja{i}", 
                        index=False,
                        header=True,
                        startrow=6,
                        startcol=2
                    )

            # 2. Declaración de hojas de ESTADOS:
            wb = load_workbook(path_xlsx)
            situaFinanciera = wb['Hoja1']
            estaresultados = wb['Hoja2']
            patrimonio = wb['Hoja3']
            flujoEfectivo = wb['Hoja4']

            # 3. ABRIR CON OPENPYXL PARA APLICAR ESTILOS MANUALES
            FormatoSituacionFinanciera(situaFinanciera, nombreEmpresa)
            FormatoResultados(estaresultados, nombreEmpresa)
            FormatoPatrimonio(wb, patrimonio, nombreEmpresa)
            FormatoFlujoEfectivo(flujoEfectivo, nombreEmpresa)

            # 4. ELIMINAR HOJAS QUE NO SON NECESARIAS
            wb.remove(wb['Hoja5'])
            wb.remove(wb['Hoja6'])

            # 5. Guardar el archivo con estilos
            wb.save(path_xlsx)

def hojas(ws):
    if ws.title == 'Hoja1':
        ws['A1'].value = "ESTADO DE SITUACIÓN FINANCIERA"
        FILAS_GRISES = [8,9,24,25,43,44,45,46,59,60,72,73,74,82,83]

    elif ws.title == 'Hoja2':
        ws['A1'].value = "ESTADO DE RESULTADOS"
        FILAS_GRISES = [8,10,16,28,32]

    elif ws.title == 'Hoja3':
        ws['A1'].value = "ESTADO DE PATRIMONIO NETO"
        ws.column_dimensions[get_column_letter(1)].width = 20
        for col in range(2, ws.max_column+1):
            ws.column_dimensions[get_column_letter(col)].width = 45
        FILAS_GRISES = [8,11,16,30,31,32,35,40,54,55,56,59,64,78,79,80,83,88,102,103,104,107,112,126,127,128,131,136,150,151]

    elif ws.title == 'Hoja4':
        ws['A1'].value = "ESTADO DE FLUJO DE EFECTIVO"
        FILAS_GRISES = [8,9,10,18,28,47,48,49,62,75,76,77,84,95]
    return FILAS_GRISES

def FormatoSituacionFinanciera(nroHoja, nombre):
    ws = nroHoja
    nombreEmpresa = nombre

    max_row = ws.max_row
    max_col = ws.max_column
    
    # Ajustar el ancho de las columnas
    ws.column_dimensions[get_column_letter(1)].width = 2
    ws.column_dimensions[get_column_letter(2)].width = 2
    ws.column_dimensions[get_column_letter(3)].width = 45
    for i in range(4, max_col + 5):
        ws.column_dimensions[get_column_letter(i)].width = 14.5

    # Título Principal (Fila 1, Columna A)
    ws['A1'].font = fuente_titulo
    ws['A3'].value = "Periodo: Anual"
    ws['A3'].font = negrita
    ws['A4'].value = f"Empresa: {nombreEmpresa}"
    ws['A4'].font = negrita
    ws['A5'].value = "Tipo: Individual"
    ws['A5'].font = negrita

    # Ajustamos la fila de los encabezados (ej. Fila 6 del Excel)
    DATA_START_ROW = 7
    FILAS_GRISES = hojas(ws)

    HEADER_ROW = 7
    for col in range(3, max_col + 5):
        cell = ws.cell(row=HEADER_ROW, column=col) 
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row in range(DATA_START_ROW, max_row + 1):
        for col in range(3, max_col + 5):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.font = Contenido

            if col == 3:
                cell.alignment = Alignment(horizontal='left', vertical='center')

            if row in FILAS_GRISES:
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.fill = SECTION_FILL

                if col in range(4, max_col + 5):
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.number_format = FORMATO_NUMERICO_FINANCIERO
            
            if col in range(4, max_col + 5):
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.number_format = FORMATO_NUMERICO_FINANCIERO
            
            if row == 9:
                cell = ws.cell(row=DATA_START_ROW, column=col) 
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='center', vertical='center')

def FormatoResultados(nroHoja, nombre):
    ws = nroHoja
    nombreEmpresa = nombre

    max_row = ws.max_row
    max_col = ws.max_column
    
    # Ajustar el ancho de las columnas
    ws.column_dimensions[get_column_letter(1)].width = 2
    ws.column_dimensions[get_column_letter(2)].width = 2
    ws.column_dimensions[get_column_letter(3)].width = 45
    for i in range(4, max_col + 5):
        ws.column_dimensions[get_column_letter(i)].width = 14.5

    # Título Principal (Fila 1, Columna A)
    ws['A1'].font = fuente_titulo
    ws['A3'].value = "Periodo: Anual"
    ws['A3'].font = negrita
    ws['A4'].value = f"Empresa: {nombreEmpresa}"
    ws['A4'].font = negrita
    ws['A5'].value = "Tipo: Individual"
    ws['A5'].font = negrita

    # Ajustamos la fila de los encabezados (ej. Fila 6 del Excel)
    DATA_START_ROW = 7
    
    FILAS_GRISES = hojas(ws)

    for row in range(DATA_START_ROW, max_row - 14):
        for col in range(3, max_col + 5):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.font = Contenido

            if col == 3:
                cell.alignment = Alignment(horizontal='left', vertical='center')

            if row in FILAS_GRISES:
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.fill = SECTION_FILL

                if col in range(4, max_col + 5):
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.number_format = FORMATO_NUMERICO_FINANCIERO
            
            if col in range(4, max_col + 5):
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.number_format = FORMATO_NUMERICO_FINANCIERO
            
            if row == 9:
                cell = ws.cell(row=DATA_START_ROW, column=col) 
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='center', vertical='center')

    limpiar_rango_Formato(ws, 'C33:I47')

def FormatoFlujoEfectivo(nroHoja, nombre):
    ws = nroHoja
    nombreEmpresa = nombre

    max_row = ws.max_row
    max_col = ws.max_column
    
    # Ajustar el ancho de las columnas
    ws.column_dimensions[get_column_letter(1)].width = 2
    ws.column_dimensions[get_column_letter(2)].width = 2
    ws.column_dimensions[get_column_letter(3)].width = 45
    for i in range(4, max_col + 5):
        ws.column_dimensions[get_column_letter(i)].width = 14.5

    # Título Principal (Fila 1, Columna A)
    ws['A1'].font = fuente_titulo
    ws['A3'].value = "Periodo: Anual"
    ws['A3'].font = negrita
    ws['A4'].value = f"Empresa: {nombreEmpresa}"
    ws['A4'].font = negrita
    ws['A5'].value = "Tipo: Individual"
    ws['A5'].font = negrita

    # Ajustamos la fila de los encabezados (ej. Fila 6 del Excel)
    DATA_START_ROW = 7
    
    FILAS_GRISES = hojas(ws)

    for row in range(DATA_START_ROW, max_row - 4):
        for col in range(3, max_col + 5):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.font = Contenido

            if col == 3:
                cell.alignment = Alignment(horizontal='left', vertical='center')

            if row in FILAS_GRISES:
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.fill = SECTION_FILL

                if col in range(4, max_col + 5):
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.number_format = FORMATO_NUMERICO_FINANCIERO
            
            if col in range(4, max_col + 5):
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.number_format = FORMATO_NUMERICO_FINANCIERO
            
            if row == 9:
                cell = ws.cell(row=DATA_START_ROW, column=col) 
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='center', vertical='center')
    limpiar_rango_Formato(ws, "C96:I100")

def FormatoPatrimonio(wb, nroHoja, nombre):
    ws = nroHoja
    nombreEmpresa = nombre

    max_row = ws.max_row
    max_col = ws.max_column

    ReordenarTabla(ws, max_col)
    
    # Ajustar el ancho de las columnas
    ws.column_dimensions[get_column_letter(4)].width = 14.5
    for i in range(5, max_col + 5):
        ws.column_dimensions[get_column_letter(i)].width = 10

    # Título Principal (Fila 1, Columna A)
    ws['A1'].font = fuente_titulo
    ws['A3'].value = "Periodo: Anual"
    ws['A3'].font = negrita
    ws['A4'].value = f"Empresa: {nombreEmpresa}"
    ws['A4'].font = negrita
    ws['A5'].value = "Tipo: Individual"
    ws['A5'].font = negrita

    # Ajustamos la fila de los encabezados (ej. Fila 6 del Excel)
    DATA_START_ROW = 7
    
    FILAS_GRISES = hojas(ws)

    for row in range(DATA_START_ROW, max_row + 97):
        for col in range(3, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.font = Contenido

            if col == 3:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

            if row in FILAS_GRISES:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.fill = SECTION_FILL

                if col in range(4, max_col + 1):
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.number_format = FORMATO_NUMERICO_FINANCIERO
            
            if col in range(4, max_col + 1):
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.number_format = FORMATO_NUMERICO_FINANCIERO
            
            if row == 9:
                cell = ws.cell(row=DATA_START_ROW, column=col) 
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws.column_dimensions[get_column_letter(1)].width = 5
    ws.column_dimensions[get_column_letter(2)].width = 5
    ws.column_dimensions[get_column_letter(3)].width = 10

    ws.merge_cells('B8:B31')    # COMBINAR Y CENTRAR - AÑO 2024
    celda2024 = ws['B8']
    celda2024.value = "2024"
    encabezadosFechasVerticales(celda2024)

    ws.merge_cells('B32:B55')    # COMBINAR Y CENTRAR - AÑO 2023
    celda2023 = ws['B32']
    celda2023.value = "2023"
    encabezadosFechasVerticales(celda2023)

    ws.merge_cells('B56:B79')    # COMBINAR Y CENTRAR - AÑO 2022
    celda2022 = ws['B56']
    celda2022.value = "2022"
    encabezadosFechasVerticales(celda2022)

    ws.merge_cells('B80:B103')    # COMBINAR Y CENTRAR - AÑO 2021
    celda2021 = ws['B80']
    celda2021.value = "2021"
    encabezadosFechasVerticales(celda2021)

    ws.merge_cells('B104:B127')    # COMBINAR Y CENTRAR - AÑO 2020
    celda2020 = ws['B104']
    celda2020.value = "2020"
    encabezadosFechasVerticales(celda2020)

def limpiar_rango_Formato(ws, rango_excel):
    min_col, min_row, max_col, max_row = range_boundaries(rango_excel)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.value = None

def encabezadosFechasVerticales(celda):
    celda.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    celda.font = Font(size=16, bold=True)
    celda.alignment = Alignment(
        horizontal='center',
        vertical='center',
        text_rotation=90  # Rota el texto 90 grados (hacia arriba)
    )
    celda.border = THIN_BORDER

def ReordenarTabla(ws, max_col):    
    # 1. Definir los límites
    # Usaremos A y la última columna (ej. E)
    MAX_COL_LETTER = get_column_letter(max_col + 1)
    MAX_ROW = ws.max_row

    RANGO_SUPERIOR = f"C8:{MAX_COL_LETTER}31"
    ws.move_range(
        RANGO_SUPERIOR,
        rows=+48,
        cols=0
    )

    RANGO_INFERIOR = f"C32:{MAX_COL_LETTER}{MAX_ROW}"
    ws.move_range(
        RANGO_INFERIOR, 
        rows=-24,
        cols=0
    )

    RANGO_SUPERIOR2 = f"C56:{MAX_COL_LETTER}79"
    ws.move_range(
        RANGO_SUPERIOR2,
        rows=-24,
        cols=0
    )

def union_archivos(path_xlsx_origen, path_xlsx_destino, columna):
    # 1. CARGAR AMBOS WORKBOOKS
    wb_origen = load_workbook(path_xlsx_origen, data_only=True)
    wb_destino = load_workbook(path_xlsx_destino)
    
    # 2. DEFINIR HOJAS DE ORIGEN Y DESTINO
    ws_Origen_hoja1 = wb_origen['Hoja1']
    ws_Destino_hoja1 = wb_destino['Hoja1']
    ws_Origen_hoja2 = wb_origen['Hoja2']
    ws_Destino_hoja2 = wb_destino['Hoja2']
    ws_Origen_hoja3 = wb_origen['Hoja3']
    ws_Destino_hoja3 = wb_destino['Hoja3']
    ws_Origen_hoja4 = wb_origen['Hoja4']
    ws_Destino_hoja4 = wb_destino['Hoja4']
    
    # 3. DEFINIR RANGO Y POSICIÓN
    rango_a_copiar_2 = 'C8:AA55'
    fila_destino_inicial = 7
    fila_destino_inicial_2 = 56
    fila_destino_inicial_3 = 104
    columna_destino_inicial = columna

    # 4. EJECUTAR COPIA
    #UNION DE SITUACION FINANCIERA - HOJA 1
    copiar_celdas(
        ws_Origen_hoja1,
        ws_Destino_hoja1,
        'D7:E83',
        fila_destino_inicial,
        columna_destino_inicial
    )

    #UNION DE RESULTADOS - HOJA 2
    copiar_celdas(
        ws_Origen_hoja2,
        ws_Destino_hoja2,
        'D7:E83',
        fila_destino_inicial,
        columna_destino_inicial
    )

    #UNION DE PATRIMONIO - HOJA 3
    """
    if columna == 5:
        copiar_celdas(
            ws_Origen_hoja3,
            ws_Destino_hoja3,
            rango_a_copiar_2,
            fila_destino_inicial_2,
            columna_destino_inicial - 3
        )
    else:
    
        copiar_celdas(
            ws_Origen_hoja3,
            ws_Destino_hoja3,
            rango_a_copiar_2,
            fila_destino_inicial_3,
            columna_destino_inicial - 5
        )
    """

    #UNION DE FLUJO DE EFECTIVOS - HOJA 4
    copiar_celdas(
        ws_Origen_hoja4,
        ws_Destino_hoja4,
        'D7:E95',
        fila_destino_inicial,
        columna_destino_inicial
    )
    
    # 5. GUARDAR EL ARCHIVO DESTINO (EL ARCHIVO ORIGEN NO SE MODIFICA)
    wb_destino.save(path_xlsx_destino)

def copiar_celdas(ws_origen, ws_destino, rango_origen: str, fila_inicio_destino: int, columna_inicio_destino: int):    
    # Obtener las coordenadas del rango de origen
    try:
        min_col, min_row, max_col, max_row = range_boundaries(rango_origen) 
    except ValueError:
        print(f"Error: Rango '{rango_origen}' no es válido.")
        return

    fila_destino = fila_inicio_destino
    
    # Iterar sobre las filas y columnas del rango de origen
    for row in ws_origen.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        col_destino = columna_inicio_destino
        
        for cell_origen in row:
            # Obtener la celda de destino
            cell_destino = ws_destino.cell(row=fila_destino, column=col_destino)
            
            # 1. Copiar Valor: Solo si la celda tiene un valor (ignora celdas combinadas secundarias)
            if cell_origen.value is not None:
                cell_destino.value = cell_origen.value 
                
                # 2. Copiar Formato Numérico
                if cell_origen.number_format:
                    cell_destino.number_format = FORMATO_NUMERICO_FINANCIERO
            else:
                # Opcional: Asegurar que la celda destino también esté vacía si el origen lo está
                cell_destino.value = None

            col_destino += 1
        fila_destino += 1

def analisis_VH(path_xlsx):
    wb = load_workbook(path_xlsx)

    try:
        Ratios = wb['Hoja5']
    except KeyError:
        Ratios = wb.copy_worksheet(wb['Hoja1'])
        Ratios.title = "Hoja5"

    Ratios['A1'].value = "RATIOS FINANCIEROS"

    sistFinan = wb['Hoja1']
    resultados = wb['Hoja2']
    flujos = wb['Hoja4']
    FormatoAnalisis1(sistFinan)
    FormatoAnalisis2(resultados)
    FormatoAnalisis3(flujos)

    wb.save(path_xlsx)

def FormatoAnalisis1(ws):
    ws.column_dimensions[get_column_letter(9)].width = 3
    ws.column_dimensions[get_column_letter(15)].width = 3
    ws.merge_cells('J6:N6')
    aplicarBorde(ws, 'J6:N83')
    ws['J6'].value = "Análisis Vertical"
    ws['J6'].fill = ENCABEZADO_NARANJA
    ws['J6'].font = negrita
    ws['J6'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('P6:S6')
    aplicarBorde(ws, 'P6:S83')
    ws['P6'].value = "Análisis Horizontal"
    ws['P6'].fill = ENCABEZADO_NARANJA
    ws['P6'].font = negrita
    ws['P6'].alignment = Alignment(horizontal='center', vertical='center')

    copiar_celdas(ws,ws,'D7:H7',7,10)
    copiar_celdas(ws,ws,'D7:G7',7,16)

    for row in range(7, 84):
        for col in range(10, 15):
            cell = ws.cell(row=row, column=col)
            if row == 7:
                cell = ws.cell(row=7, column=col) 
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            elif row in [8,9,24,25,43,44,45,46,59,60,72,73,74,82,83]:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.fill = SECTION_FILL
            
            if row in range(8, 84):
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for col in range(16, 20):
            cell = ws.cell(row=row, column=col)
            if row == 7:
                cell = ws.cell(row=7, column=col) 
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            elif row in [8,9,24,25,43,44,45,46,59,60,72,73,74,82,83]:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.fill = SECTION_FILL
            
            if row in range(8, 84):
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def FormatoAnalisis2(ws):
    ws.column_dimensions[get_column_letter(9)].width = 3
    ws.column_dimensions[get_column_letter(15)].width = 3
    ws.merge_cells('J6:N6')
    aplicarBorde(ws, 'J6:N32')
    ws['J6'].value = "Análisis Vertical"
    ws['J6'].fill = ENCABEZADO_NARANJA
    ws['J6'].font = negrita
    ws['J6'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('P6:S6')
    aplicarBorde(ws, 'P6:S32')
    ws['P6'].value = "Análisis Horizontal"
    ws['P6'].fill = ENCABEZADO_NARANJA
    ws['P6'].font = negrita
    ws['P6'].alignment = Alignment(horizontal='center', vertical='center')

    copiar_celdas(ws,ws,'D7:H7',7,10)
    copiar_celdas(ws,ws,'D7:G7',7,16)

    for row in range(7, 33):
        for col in range(10, 15):
            cell = ws.cell(row=row, column=col)
            if row == 7:
                cell = ws.cell(row=7, column=col) 
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            elif row in [8,10,16,28,32]:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.fill = SECTION_FILL
            
            if row in range(8, 33):
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for col in range(16, 20):
            cell = ws.cell(row=row, column=col)
            if row == 7:
                cell = ws.cell(row=7, column=col) 
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            elif row in [8,10,16,28,32]:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.fill = SECTION_FILL
            
            if row in range(8, 33):
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def FormatoAnalisis3(ws):
    ws.column_dimensions[get_column_letter(9)].width = 3
    ws.column_dimensions[get_column_letter(15)].width = 3
    ws.merge_cells('J6:N6')
    aplicarBorde(ws, 'J6:N95')
    ws['J6'].value = "Análisis Vertical"
    ws['J6'].fill = ENCABEZADO_NARANJA
    ws['J6'].font = negrita
    ws['J6'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('P6:S6')
    aplicarBorde(ws, 'P6:S95')
    ws['P6'].value = "Análisis Horizontal"
    ws['P6'].fill = ENCABEZADO_NARANJA
    ws['P6'].font = negrita
    ws['P6'].alignment = Alignment(horizontal='center', vertical='center')

    copiar_celdas(ws,ws,'D7:H7',7,10)
    copiar_celdas(ws,ws,'D7:G7',7,16)

    for row in range(7, 96):
        for col in range(10, 15):
            cell = ws.cell(row=row, column=col)
            if row == 7:
                cell = ws.cell(row=7, column=col) 
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            elif row in [8,9,10,18,28,47,48,49,62,75,76,77,84,95]:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.fill = SECTION_FILL
        
        for col in range(16, 20):
            cell = ws.cell(row=row, column=col)
            if row == 7:
                cell = ws.cell(row=7, column=col) 
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            elif row in [8,9,10,18,28,47,48,49,62,75,76,77,84,95]:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.fill = SECTION_FILL

def limpiar_rango_Libre(ws, rango_excel):
    BORDE_POR_DEFECTO = Border(left=Side(style=None), 
                               right=Side(style=None), 
                               top=Side(style=None), 
                               bottom=Side(style=None))
    try:
        min_col, min_row, max_col, max_row = range_boundaries(rango_excel)
    except Exception:
        print(f"Error: El rango '{rango_excel}' no es un formato de rango válido.")
        return

    # 3. Iterar y limpiar cada celda
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
            cell.border = BORDE_POR_DEFECTO
            cell.number_format = 'General'
            
def aplicarBorde(ws, rango_excel):
    min_col, min_row, max_col, max_row = range_boundaries(rango_excel)

    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER

def analisis_Ratios(path_xlsx):
    wb = load_workbook(path_xlsx)
    ws = wb['Hoja5']

    ws.column_dimensions[get_column_letter(9)].width = 3
    ws.column_dimensions[get_column_letter(10)].width = 30
    copiar_celdas(ws,ws,'D7:H7',7,11)
    copiar_celdas(ws,ws,'D7:H7',11,11)
    copiar_celdas(ws,ws,'D7:H7',16,11)
    copiar_celdas(ws,ws,'D7:H7',20,11)

    ws['J7'].value = "RATIOS DE LIQUIDEZ"
    ws['J7'].fill = ENCABEZADO_NARANJA
    ws['J7'].font = Font(size=11, bold=True)
    ws['J8'].value = "Liquidez Corriente"
    ws['J9'].value = "Prueba Ácida"

    ws['J11'].value = "RATIOS DE GESTIÓN"
    ws['J11'].fill = ENCABEZADO_NARANJA
    ws['J11'].font = Font(size=11, bold=True)
    ws['J12'].value = "Rotación de Cuentas por cobrar"
    ws['J13'].value = "Rotación de Inventarios"
    ws['J14'].value = "Rotación de Activos Totales"

    ws['J16'].value = "RATIOS DE ENDEUDAMIENTO"
    ws['J16'].fill = ENCABEZADO_NARANJA
    ws['J16'].font = Font(size=11, bold=True)
    ws['J17'].value = "Razón de deuda total"
    ws['J18'].value = "Razón de deuda/patrimonio"

    ws['J20'].value = "RATIOS DE RENTABILIDAD"
    ws['J20'].fill = ENCABEZADO_NARANJA
    ws['J20'].font = Font(size=11, bold=True)
    ws['J21'].value = "Margen neto"
    ws['J22'].value = "ROA"
    ws['J23'].value = "ROE"

    aplicarBorde(ws, 'J7:O9')
    aplicarBorde(ws, 'J11:O14')
    aplicarBorde(ws, 'J16:O18')
    aplicarBorde(ws, 'J20:O23')

    for row in range(6, 24):
        for col in range(11, 16):
            cell = ws.cell(row=row, column=col)
            if row in [7,11,16,20]:
                cell = ws.cell(row=row, column=col) 
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.number_format = FORMATO_NUMERICO
            elif row in [8,9,11,12,13]:
                cell.number_format = FORMATO_NUMERICO_FINANCIERO
            else:
                cell.number_format = FORMATO_PORCENTAJE_DOS_DECIMALES
    wb.save(path_xlsx)

def graficosRatios(path_xlsx):
    wb = load_workbook(path_xlsx)
    try:
        GraRati = wb['Hoja6']
    except KeyError:
        GraRati = wb.create_sheet(title="Hoja6", index=None)
    ws = GraRati

    dir_path = os.path.dirname(path_xlsx)
    nombre = os.path.basename(dir_path)
    nombreEmpresa = nombre.replace('_', ' ')

    ws['A1'].value = "GRÁFICOS DE RATIOS"
    ws['A1'].font = fuente_titulo
    ws['A3'].value = "Periodo: Anual"
    ws['A3'].font = negrita
    ws['A4'].value = f"Empresa: {nombreEmpresa}"
    ws['A4'].font = negrita
    ws['A5'].value = "Tipo: Individual"
    ws['A5'].font = negrita

    ws['C7'].value = "RATIOS DE LIQUIDEZ"

    ws.column_dimensions[get_column_letter(1)].width = 3
    ws.column_dimensions[get_column_letter(2)].width = 3
    for i in range(3,30):
        ws.column_dimensions[get_column_letter(i)].width = 15
    for i in [5,8,11,14]:
        ws.column_dimensions[get_column_letter(i)].width = 32

    ws.merge_cells('C7:G7')
    ws.merge_cells('E8:E14')
    ws['C7'].value = "RATIOS DE LIQUIDEZ"
    ws['C7'].fill = ENCABEZADO_PURPURA
    ws['C7'].font = HEADER_FONT
    ws['C7'].font = Font(size=12, bold=True)
    ws.merge_cells('C8:D8')
    ws['C8'].value = "Liquidez Corriente"
    ws['C8'].fill = ENCABEZADO_NARANJA
    ws['C9'].value = "Año"
    ws['C9'].fill =ENCABEZADO_CELESTE
    ws['D9'].value = "Valor"
    ws['D9'].fill =ENCABEZADO_CELESTE
    ws.merge_cells('F8:G8')
    ws['F8'].value = "Prueba Ácida"
    ws['F8'].fill = ENCABEZADO_NARANJA
    ws['F9'].value = "Año"
    ws['F9'].fill =ENCABEZADO_CELESTE
    ws['G9'].value = "Valor"
    ws['G9'].fill =ENCABEZADO_CELESTE

    ws.merge_cells('I7:P7')
    ws.merge_cells('K8:K13')
    ws.merge_cells('N8:N13')
    ws['I7'].value = "RATIOS DE GESTIÓN"
    ws['I7'].fill = ENCABEZADO_PURPURA
    ws['I7'].font = HEADER_FONT
    ws['I7'].font = Font(size=12, bold=True)
    ws.merge_cells('I8:J8')
    ws['I8'].value = "Rotación de Cuentas por cobrar"
    ws['I8'].fill = ENCABEZADO_NARANJA
    ws['I9'].value = "Año"
    ws['I9'].fill =ENCABEZADO_CELESTE
    ws['J9'].value = "Valor"
    ws['J9'].fill =ENCABEZADO_CELESTE
    ws.merge_cells('L8:M8')
    ws['L8'].value = "Rotación de Inventarios"
    ws['L8'].fill = ENCABEZADO_NARANJA
    ws['L9'].value = "Año"
    ws['L9'].fill =ENCABEZADO_CELESTE
    ws['M9'].value = "Valor"
    ws['M9'].fill =ENCABEZADO_CELESTE
    ws.merge_cells('O8:P8')
    ws['O8'].value = "Rotación de Activos Totales"
    ws['O8'].fill = ENCABEZADO_NARANJA
    ws['O9'].value = "Año"
    ws['O9'].fill = ENCABEZADO_CELESTE
    ws['P9'].value = "Valor"
    ws['P9'].fill = ENCABEZADO_CELESTE

    ws.merge_cells('C30:G30')
    ws.merge_cells('E31:E37')
    ws['C30'].value = "RATIOS DE ENDEUDAMIENTO"
    ws['C30'].fill = ENCABEZADO_PURPURA
    ws['C30'].font = HEADER_FONT
    ws['C30'].font = Font(size=12, bold=True)
    ws.merge_cells('C31:D31')
    ws['C31'].value = "Razón de deuda total"
    ws['C31'].fill = ENCABEZADO_NARANJA
    ws['C32'].value = "Año"
    ws['C32'].fill = ENCABEZADO_CELESTE
    ws['D32'].value = "Valor"
    ws['D32'].fill = ENCABEZADO_CELESTE
    ws.merge_cells('F31:G31')
    ws['F31'].value = "Razón de deuda/patrimonio"
    ws['F31'].fill = ENCABEZADO_NARANJA
    ws['F32'].value = "Año"
    ws['F32'].fill = ENCABEZADO_CELESTE
    ws['G32'].value = "Valor"
    ws['G32'].fill = ENCABEZADO_CELESTE

    ws.merge_cells('I30:P30')
    ws.merge_cells('K31:K37')
    ws.merge_cells('N31:N37')
    ws['I30'].value = "RATIOS DE RENTABILIDAD"
    ws['I30'].fill = ENCABEZADO_PURPURA
    ws['I30'].font = HEADER_FONT
    ws['I30'].font = Font(size=12, bold=True)
    ws.merge_cells('I31:J31')
    ws['I31'].value = "Margen neto"
    ws['I31'].fill = ENCABEZADO_NARANJA
    ws['I32'].value = "Año"
    ws['I32'].fill = ENCABEZADO_CELESTE
    ws['J32'].value = "Valor"
    ws['J32'].fill = ENCABEZADO_CELESTE
    ws.merge_cells('L31:M31')
    ws['L31'].value = "ROA"
    ws['L31'].fill = ENCABEZADO_NARANJA
    ws['L32'].value = "Año"
    ws['L32'].fill = ENCABEZADO_CELESTE
    ws['M32'].value = "Valor"
    ws['M32'].fill = ENCABEZADO_CELESTE
    ws.merge_cells('O31:P31')
    ws['O31'].value = "ROE"
    ws['O31'].fill = ENCABEZADO_NARANJA
    ws['O32'].value = "Año"
    ws['O32'].fill = ENCABEZADO_CELESTE
    ws['P32'].value = "Valor"
    ws['P32'].fill = ENCABEZADO_CELESTE

    aplicarBorde(ws, 'C7:G14')
    centrar_rango(ws, 'C7:G14')
    aplicarBorde(ws, 'I7:P13')
    centrar_rango(ws, 'I7:P13')
    aplicarBorde(ws, 'C30:G37')
    centrar_rango(ws, 'C30:G37')
    aplicarBorde(ws, 'I30:P37')
    centrar_rango(ws, 'I30:P37')

    wb.save(path_xlsx)
    
def renombrar(path_xlsx):
    wb = load_workbook(path_xlsx)
    sistFinan = wb['Hoja1']
    resultados = wb['Hoja2']
    patrimonio = wb['Hoja3']
    flujos = wb['Hoja4']
    ratios = wb['Hoja5']
    graratios = wb['Hoja6']

    sistFinan.title = "Estado de Situación Financiera"
    resultados.title = "Estado de Resultados"
    patrimonio.title = "Estado de Patrimonio Neto"
    flujos.title = "Estado de Flujo de Efectivo"
    ratios.title = "Ratios Financieros"
    graratios.title = "Gráficos de Ratios Financieros"
    wb.save(path_xlsx)

def centrar_rango(ws, rango):
    alineacion_centrada = Alignment(horizontal='center', vertical='center')
    min_col, min_row, max_col, max_row = range_boundaries(rango)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = alineacion_centrada

def valor(wb, nombre_hoja, celda):
    valor = wb[nombre_hoja][celda].value
    if valor is None:
        return "0"
    s = str(valor).strip()
    # Manejar formato contable con paréntesis (negativos)
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]

    # Normalizar separadores para convertir a número
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")

    try:
        num = float(s)
        # Formatear salida con punto miles y coma decimal
        if num.is_integer():
            return f"{int(num):,}".replace(",", ".")  # enteros sin decimales
        else:
            return f"{num:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except ValueError:
        return "0"

def analisisVertical(path_xlsx):
    wb = load_workbook(path_xlsx)
    ws1 = wb['Hoja1']
    ws2 = wb['Hoja2']
    ws4 = wb['Hoja4']

    for row in range(8, 84):    # SITUACIÓN FINANCIERA
        for offset, col in enumerate(range(4, 9)):
            num = ws1.cell(row=row, column=col).value
            num = convertir_a_numero(num)
            den = ws1.cell(row=44, column=col).value
            den = convertir_a_numero(den)

            if den not in (0, None):
                resultado = num / den
            else:
                resultado = 0.00

            celda_resultado = ws1.cell(row=row, column=10 + offset, value=resultado)
            celda_resultado.value = resultado
            celda_resultado.number_format = FORMATO_PORCENTAJE_DOS_DECIMALES


    for row in range(8, 33):    # RESULTADOS
        for offset, col in enumerate(range(4, 9)):
            num = ws2.cell(row=row, column=col).value
            num = convertir_a_numero(num)
            den = ws2.cell(row=8, column=col).value
            den = convertir_a_numero(den)

            if den not in (0, None):
                resultado = num / den
            else:
                resultado = 0.00

            celda_resultado = ws2.cell(row=row, column=10 + offset, value=resultado)
            celda_resultado.value = resultado
            celda_resultado.number_format = FORMATO_PORCENTAJE_DOS_DECIMALES


    for row in range(8, 48):    # FLUJO DE EFECTIVO
        for offset, col in enumerate(range(4, 9)):
            num = ws4.cell(row=row, column=col).value
            num = convertir_a_numero(num)
            den = ws4.cell(row=47, column=col).value
            den = convertir_a_numero(den)

            if den not in (0, None):
                resultado = num / den
            else:
                resultado = 0.00

            celda_resultado = ws4.cell(row=row, column=10 + offset, value=resultado)
            celda_resultado.value = resultado
            celda_resultado.number_format = FORMATO_PORCENTAJE_DOS_DECIMALES

    for row in range(48, 76):    # FLUJO DE EFECTIVO
        for offset, col in enumerate(range(4, 9)):
            num = ws4.cell(row=row, column=col).value
            num = convertir_a_numero(num)
            den = ws4.cell(row=75, column=col).value
            den = convertir_a_numero(den)

            if den not in (0, None):
                resultado = num / den
            else:
                resultado = 0.00

            celda_resultado = ws4.cell(row=row, column=10 + offset, value=resultado)
            celda_resultado.value = resultado
            celda_resultado.number_format = FORMATO_PORCENTAJE_DOS_DECIMALES


    for row in range(77, 96):    # FLUJO DE EFECTIVO
        for offset, col in enumerate(range(4, 9)):
            num = ws4.cell(row=row, column=col).value
            num = convertir_a_numero(num)
            den = ws4.cell(row=95, column=col).value
            den = convertir_a_numero(den)

            if den not in (0, None):
                resultado = num / den
            else:
                resultado = 0.00

            celda_resultado = ws4.cell(row=row, column=10 + offset, value=resultado)
            celda_resultado.value = resultado
            celda_resultado.number_format = FORMATO_PORCENTAJE_DOS_DECIMALES

    wb.save(path_xlsx)

def analisisHorizontal(path_xlsx):
    wb = load_workbook(path_xlsx)
    ws1 = wb['Hoja1']
    ws2 = wb['Hoja2']
    ws4 = wb['Hoja4']

    for row in range(8, 84):    # SITUACIÓN FINANCIERA
        for offset, col in enumerate(range(4, 8)):
            num = ws1.cell(row=row, column=col).value
            num = convertir_a_numero(num)
            den = ws1.cell(row=row, column=col+1).value
            den = convertir_a_numero(den)

            if den not in (0, None):
                resultado = (num - den) / den
            else:
                resultado = 0.00

            celda_resultado = ws1.cell(row=row, column=16 + offset, value=resultado)
            celda_resultado.value = resultado
            celda_resultado.number_format = FORMATO_PORCENTAJE_DOS_DECIMALES

    for row in range(8, 33):    # RESULTADOS
        for offset, col in enumerate(range(4, 8)):
            num = ws2.cell(row=row, column=col).value
            num = convertir_a_numero(num)
            den = ws2.cell(row=row, column=col+1).value
            den = convertir_a_numero(den)

            if den not in (0, None):
                resultado = (num - den) / den
            else:
                resultado = 0.00

            celda_resultado = ws2.cell(row=row, column=16 + offset, value=resultado)
            celda_resultado.value = resultado
            celda_resultado.number_format = FORMATO_PORCENTAJE_DOS_DECIMALES

    for row in range(8, 96):    # FLUJO DE EFECTIVO
        for offset, col in enumerate(range(4, 8)):
            num = ws4.cell(row=row, column=col).value
            num = convertir_a_numero(num)
            den = ws4.cell(row=row, column=col+1).value
            den = convertir_a_numero(den)

            if den not in (0, None):
                resultado = (num - den) / den
            else:
                resultado = 0.00

            celda_resultado = ws4.cell(row=row, column=16 + offset, value=resultado)
            celda_resultado.value = resultado
            celda_resultado.number_format = FORMATO_PORCENTAJE_DOS_DECIMALES

    wb.save(path_xlsx)

def analisisRatiosCalculo(path_xlsx):
    wb = load_workbook(path_xlsx)
    ws1 = wb['Hoja5']
    ws2 = wb['Hoja2']

    for offset, col in enumerate(range(4, 9)):  # LIQUIDEZ CORRIENTE
        num = ws1.cell(row=24, column=col).value
        num = convertir_a_numero(num)
        den = ws1.cell(row=59, column=col).value
        den = convertir_a_numero(den)

        if den not in (0, None):
            resultado = num / den
        else:
            resultado = 0.00

        celda_resultado = ws1.cell(row=8, column=11 + offset, value=resultado)
        celda_resultado.value = resultado
        celda_resultado.number_format = FORMATO_NUMERICO_FINANCIERO


    for offset, col in enumerate(range(4, 9)):  # PRUEBA ÁCIDA
        num1 = ws1.cell(row=24, column=col).value
        num1 = convertir_a_numero(num1)
        num2 = ws1.cell(row=17, column=col).value
        num2 = convertir_a_numero(num2)
        den = ws1.cell(row=59, column=col).value
        den = convertir_a_numero(den)

        if den not in (0, None):
            resultado = (num1 - num2) / den
        else:
            resultado = 0.00

        celda_resultado = ws1.cell(row=9, column=11 + offset, value=resultado)
        celda_resultado.value = resultado
        celda_resultado.number_format = FORMATO_NUMERICO_FINANCIERO

    
    for offset, col in enumerate(range(4, 9)):  # RAZÓN DE DEUDA TOTAL
        num = ws1.cell(row=73, column=col).value
        num = convertir_a_numero(num)
        den = ws1.cell(row=44, column=col).value
        den = convertir_a_numero(den)

        if den not in (0, None):
            resultado = num / den
        else:
            resultado = 0.00

        celda_resultado = ws1.cell(row=17, column=11 + offset, value=resultado)
        celda_resultado.value = resultado
        celda_resultado.number_format = FORMATO_PORCENTAJE_DOS_DECIMALES


    for offset, col in enumerate(range(4, 9)):  # RAZÓN DE DEUDA/PATRIMONIO
        num = ws1.cell(row=73, column=col).value
        num = convertir_a_numero(num)
        den = ws1.cell(row=82, column=col).value
        den = convertir_a_numero(den)

        if den not in (0, None):
            resultado = num / den
        else:
            resultado = 0.00

        celda_resultado = ws1.cell(row=18, column=11 + offset, value=resultado)
        celda_resultado.value = resultado
        celda_resultado.number_format = FORMATO_PORCENTAJE_DOS_DECIMALES


    for offset, col in enumerate(range(4, 9)):  # MARGEN NETO
        num = ws2.cell(row=32, column=col).value
        num = convertir_a_numero(num)
        den = ws2.cell(row=8, column=col).value
        den = convertir_a_numero(den)

        if den not in (0, None):
            resultado = num / den
        else:
            resultado = 0.00

        celda_resultado = ws1.cell(row=21, column=11 + offset, value=resultado)
        celda_resultado.value = resultado
        celda_resultado.number_format = FORMATO_PORCENTAJE_DOS_DECIMALES


    for offset, col in enumerate(range(4, 9)):  # ROA
        num = ws2.cell(row=32, column=col).value
        num = convertir_a_numero(num)
        den = ws1.cell(row=44, column=col).value
        den = convertir_a_numero(den)

        if den not in (0, None):
            resultado = num / den
        else:
            resultado = 0.00

        celda_resultado = ws1.cell(row=22, column=11 + offset, value=resultado)
        celda_resultado.value = resultado
        celda_resultado.number_format = FORMATO_PORCENTAJE_DOS_DECIMALES


    for offset, col in enumerate(range(4, 9)):  # ROE
        num = ws2.cell(row=32, column=col).value
        num = convertir_a_numero(num)
        den = ws1.cell(row=82, column=col).value
        den = convertir_a_numero(den)

        if den not in (0, None):
            resultado = num / den
        else:
            resultado = 0.00

        celda_resultado = ws1.cell(row=23, column=11 + offset, value=resultado)
        celda_resultado.value = resultado
        celda_resultado.number_format = FORMATO_PORCENTAJE_DOS_DECIMALES


    for offset, col in enumerate(range(4, 8)):  # ROTACIÓN DE CUENTAS POR COBRAR
        num = ws2.cell(row=8, column=col).value
        num = convertir_a_numero(num)

        # Valores columna col (D-G)
        d1 = convertir_a_numero(ws1.cell(row=12, column=col).value)
        d2 = convertir_a_numero(ws1.cell(row=13, column=col).value)
        d3 = convertir_a_numero(ws1.cell(row=14, column=col).value)
        d4 = convertir_a_numero(ws1.cell(row=15, column=col).value)
        d5 = convertir_a_numero(ws1.cell(row=28, column=col).value)
        d6 = convertir_a_numero(ws1.cell(row=29, column=col).value)
        d7 = convertir_a_numero(ws1.cell(row=30, column=col).value)
        d8 = convertir_a_numero(ws1.cell(row=31, column=col).value)

        # Valores columna col+1 (E-H)
        d9  = convertir_a_numero(ws1.cell(row=12, column=col+1).value)
        d10 = convertir_a_numero(ws1.cell(row=13, column=col+1).value)
        d11 = convertir_a_numero(ws1.cell(row=14, column=col+1).value)
        d12 = convertir_a_numero(ws1.cell(row=15, column=col+1).value)
        d13 = convertir_a_numero(ws1.cell(row=28, column=col+1).value)
        d14 = convertir_a_numero(ws1.cell(row=29, column=col+1).value)
        d15 = convertir_a_numero(ws1.cell(row=30, column=col+1).value)
        d16 = convertir_a_numero(ws1.cell(row=31, column=col+1).value)

        den = (d1+d2+d3+d4+d5+d6+d7+d8 + d9+d10+d11+d12+d13+d14+d15+d16) / 2

        if den != 0:
            resultado = num / den
        else:
            resultado = 0.00

        celda_resultado = ws1.cell(row=12, column=11 + offset, value=resultado)
        celda_resultado.value = resultado
        celda_resultado.number_format = FORMATO_NUMERICO_FINANCIERO
    ws1['O12'].value = 0.00


    for offset, col in enumerate(range(4, 8)):  # ROTACIÓN DE INVENTARIOS
        num = ws2.cell(row=9, column=col).value
        num = convertir_a_numero(num)

        d1 = convertir_a_numero(ws1.cell(row=17, column=col).value)
        d2 = convertir_a_numero(ws1.cell(row=33, column=col+1).value)
        d3 = convertir_a_numero(ws1.cell(row=17, column=col).value)
        d4 = convertir_a_numero(ws1.cell(row=33, column=col+1).value)

        den = (d1+d2+d3+d4) / 2

        if den != 0:
            resultado = (num / den) * -1
        else:
            resultado = 0.00

        celda_resultado = ws1.cell(row=13, column=11 + offset, value=resultado)
        celda_resultado.value = resultado
        celda_resultado.number_format = FORMATO_NUMERICO_FINANCIERO
    ws1['O13'].value = 0.00


    for offset, col in enumerate(range(4, 9)):  # ROTACIÓN DE ACTIVOS TOTALES
        num = ws2.cell(row=8, column=col).value
        num = convertir_a_numero(num)

        den = convertir_a_numero(ws1.cell(row=44, column=col).value)

        if den != 0:
            resultado = num / den
        else:
            resultado = 0.00

        celda_resultado = ws1.cell(row=14, column=11 + offset, value=resultado)
        celda_resultado.value = resultado
        celda_resultado.number_format = FORMATO_NUMERICO_FINANCIERO


    

    wb.save(path_xlsx)

def valor(wb, nombre_hoja, celda):
    valor = wb[nombre_hoja][celda].value
    return convertir_a_numero(valor)

def convertir_a_numero(valor): 
    if valor is None: 
        return 0.0 
    s = str(valor).strip() 

    # Manejar paréntesis (negativos contables) 

    if s.startswith("(") and s.endswith(")"): 
        s = "-" + s[1:-1].strip() 
    
    # Eliminar separadores de miles (,) y mantener el punto como decimal
    s = s.replace(",", "") 
    try: 
        return float(s) 
    except ValueError: 
        return 0.0
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Color
import matplotlib
matplotlib.use('Agg')  # ¡IMPORTANTE! Usar backend no interactivo ANTES de importar pyplot
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl.drawing.image import Image
import matplotlib
matplotlib.use('Agg')  # Backend no interactivo
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl.drawing.image import Image
import os
import tempfile
import io 
from openpyxl.utils.cell import range_boundaries, coordinate_from_string  





# Color de relleno de encabezado (Azul Oscuro/Teal)
HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")

# Color de relleno de encabezado (Azul Oscuro/Teal)
ENCABEZADO_NARANJA = PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid")

ENCABEZADO_PURPURA = PatternFill(start_color="BE78F8", end_color="BE78F8", fill_type="solid")

ENCABEZADO_CELESTE = PatternFill(start_color="65FFD7", end_color="65FFD7", fill_type="solid")

# Color de relleno para la sección principal (Gris Claro)
SECTION_FILL = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

# Fuente para encabezados (Blanco y Negrita)
HEADER_FONT = Font(size=9, bold=True, color="FFFFFF")
Contenido = Font(name="Franklin Gothic Medium Cond", size=9)

# Estilo de borde (borde fino para todas las celdas de datos)
THIN_SIDE = Side(style='thin', color="000000")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

fuente_titulo = Font(size=13.5, bold=True)
negrita = Font(bold=True)
cuentas = Alignment(horizontal='center', vertical='center')

# Formato de SIMBOLO DECIMAL (.), MILES (,) SIMBOLO NEGATIVO (-)
FORMATO_NUMERICO_FINANCIERO = '#,##0.00;-#,##0.00'
FORMATO_NUMERICO = '#,##0;-#,##0.'
FORMATO_PORCENTAJE_DOS_DECIMALES = '0.00%'

# -----------------------------
def formato_xls_xlsx(CARPETA_FINANCIEROS):
    #FORMATO DE LOS ESTADOS FINANCIEROS - Ultimos 5 años
    for archivo in os.listdir(CARPETA_FINANCIEROS):
        if archivo.endswith(".xls"):
            path_xls = os.path.join(CARPETA_FINANCIEROS, archivo)
            path_xlsx = path_xls + "x"
            dir_path = os.path.dirname(path_xls)    #Ruta del excel
            nombre = os.path.basename(dir_path)    #Nombre de la empresa
            nombreEmpresa = nombre.replace('_', ' ')
            tablas = pd.read_html(path_xls)

            # 1. GUARDAR DATAFRAME
            with pd.ExcelWriter(path_xlsx, engine="openpyxl") as writer:
                for i, df in enumerate(tablas, start=1):

                    if df.shape[1] >= 4 and i!=3:   #TODOS MENOS LA DE PATRIMONIO O FIRMANTES ELIMINAN LA COLUMNA B
                        df_cleaned = df.drop(df.columns[[1, 3]], axis=1, errors='ignore')
                    else:
                        df_cleaned = df

                    def clean_and_coerce(series):
                        s = series.astype(str)
                        s = s.str.replace('(', '-', regex=False).str.replace(')', '', regex=False)
                        s = s.str.replace(',', '', regex=False)
                        return pd.to_numeric(s, errors='coerce')

                    if df_cleaned.shape[1] >= 3 and (i!=3 and i!=6):
                        df_cleaned.iloc[:, [1, 2]] = df_cleaned.iloc[:, [1, 2]].apply(clean_and_coerce)
                    else:
                        df_cleaned.iloc[:, 2:df_cleaned.shape[1]] = df_cleaned.iloc[:, 2:df_cleaned.shape[1]].apply(clean_and_coerce)

                    df_cleaned.to_excel(
                        writer, 
                        sheet_name=f"Hoja{i}", 
                        index=False,
                        header=True,
                        startrow=6,
                        startcol=2
                    )

            # 2. Declaración de hojas de ESTADOS:
            wb = load_workbook(path_xlsx)
            situaFinanciera = wb['Hoja1']
            estaresultados = wb['Hoja2']
            patrimonio = wb['Hoja3']
            flujoEfectivo = wb['Hoja4']

            # 3. ABRIR CON OPENPYXL PARA APLICAR ESTILOS MANUALES
            FormatoSituacionFinanciera(situaFinanciera, nombreEmpresa)
            FormatoResultados(estaresultados, nombreEmpresa)
            FormatoPatrimonio(wb, patrimonio, nombreEmpresa)
            FormatoFlujoEfectivo(flujoEfectivo, nombreEmpresa)

            # 4. ELIMINAR HOJAS QUE NO SON NECESARIAS
            wb.remove(wb['Hoja5'])
            wb.remove(wb['Hoja6'])

            # 5. Guardar el archivo con estilos
            wb.save(path_xlsx)

def hojas(ws):
    if ws.title == 'Hoja1':
        ws['A1'].value = "ESTADO DE SITUACIÓN FINANCIERA"
        FILAS_GRISES = [8,9,24,25,43,44,45,46,59,60,72,73,74,82,83]

    elif ws.title == 'Hoja2':
        ws['A1'].value = "ESTADO DE RESULTADOS"
        FILAS_GRISES = [8,10,16,28,32]

    elif ws.title == 'Hoja3':
        ws['A1'].value = "ESTADO DE PATRIMONIO NETO"
        ws.column_dimensions[get_column_letter(1)].width = 20
        for col in range(2, ws.max_column+1):
            ws.column_dimensions[get_column_letter(col)].width = 45
        FILAS_GRISES = [8,11,16,30,31,32,35,40,54,55,56,59,64,78,79,80,83,88,102,103,104,107,112,126,127,128,131,136,150,151]

    elif ws.title == 'Hoja4':
        ws['A1'].value = "ESTADO DE FLUJO DE EFECTIVO"
        FILAS_GRISES = [8,9,15,28,29,30,43,56,57,58,65,76]
    return FILAS_GRISES

def FormatoSituacionFinanciera(nroHoja, nombre):
    ws = nroHoja
    nombreEmpresa = nombre

    max_row = ws.max_row
    max_col = ws.max_column
    
    # Ajustar el ancho de las columnas
    ws.column_dimensions[get_column_letter(1)].width = 2
    ws.column_dimensions[get_column_letter(2)].width = 2
    ws.column_dimensions[get_column_letter(3)].width = 45
    for i in range(4, max_col + 5):
        ws.column_dimensions[get_column_letter(i)].width = 14.5

    # Título Principal (Fila 1, Columna A)
    ws['A1'].font = fuente_titulo
    ws['A3'].value = "Periodo: Anual"
    ws['A3'].font = negrita
    ws['A4'].value = f"Empresa: {nombreEmpresa}"
    ws['A4'].font = negrita
    ws['A5'].value = "Tipo: Individual"
    ws['A5'].font = negrita

    # Ajustamos la fila de los encabezados (ej. Fila 6 del Excel)
    DATA_START_ROW = 7
    FILAS_GRISES = hojas(ws)

    HEADER_ROW = 7
    for col in range(3, max_col + 5):
        cell = ws.cell(row=HEADER_ROW, column=col) 
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row in range(DATA_START_ROW, max_row + 1):
        for col in range(3, max_col + 5):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.font = Contenido

            if col == 3:
                cell.alignment = Alignment(horizontal='left', vertical='center')

            if row in FILAS_GRISES:
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.fill = SECTION_FILL

                if col in range(4, max_col + 5):
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.number_format = FORMATO_NUMERICO_FINANCIERO
            
            if col in range(4, max_col + 5):
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.number_format = FORMATO_NUMERICO_FINANCIERO
            
            if row == 9:
                cell = ws.cell(row=DATA_START_ROW, column=col) 
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='center', vertical='center')

def FormatoResultados(nroHoja, nombre):
    ws = nroHoja
    nombreEmpresa = nombre

    max_row = ws.max_row
    max_col = ws.max_column
    
    # Ajustar el ancho de las columnas
    ws.column_dimensions[get_column_letter(1)].width = 2
    ws.column_dimensions[get_column_letter(2)].width = 2
    ws.column_dimensions[get_column_letter(3)].width = 45
    for i in range(4, max_col + 5):
        ws.column_dimensions[get_column_letter(i)].width = 14.5

    # Título Principal (Fila 1, Columna A)
    ws['A1'].font = fuente_titulo
    ws['A3'].value = "Periodo: Anual"
    ws['A3'].font = negrita
    ws['A4'].value = f"Empresa: {nombreEmpresa}"
    ws['A4'].font = negrita
    ws['A5'].value = "Tipo: Individual"
    ws['A5'].font = negrita

    # Ajustamos la fila de los encabezados (ej. Fila 6 del Excel)
    DATA_START_ROW = 7
    
    FILAS_GRISES = hojas(ws)

    for row in range(DATA_START_ROW, max_row + 1):
        for col in range(3, max_col + 5):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.font = Contenido

            if col == 3:
                cell.alignment = Alignment(horizontal='left', vertical='center')

            if row in FILAS_GRISES:
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.fill = SECTION_FILL

                if col in range(4, max_col + 5):
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.number_format = FORMATO_NUMERICO_FINANCIERO
            
            if col in range(4, max_col + 5):
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.number_format = FORMATO_NUMERICO_FINANCIERO
            
            if row == 9:
                cell = ws.cell(row=DATA_START_ROW, column=col) 
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='center', vertical='center')


def FormatoFlujoEfectivo(nroHoja, nombre):
    ws = nroHoja
    nombreEmpresa = nombre

    max_row = ws.max_row
    max_col = ws.max_column
    
    # Ajustar el ancho de las columnas
    ws.column_dimensions[get_column_letter(1)].width = 2
    ws.column_dimensions[get_column_letter(2)].width = 2
    ws.column_dimensions[get_column_letter(3)].width = 45
    for i in range(4, max_col + 5):
        ws.column_dimensions[get_column_letter(i)].width = 14.5

    # Título Principal (Fila 1, Columna A)
    ws['A1'].font = fuente_titulo
    ws['A3'].value = "Periodo: Anual"
    ws['A3'].font = negrita
    ws['A4'].value = f"Empresa: {nombreEmpresa}"
    ws['A4'].font = negrita
    ws['A5'].value = "Tipo: Individual"
    ws['A5'].font = negrita

    # Ajustamos la fila de los encabezados (ej. Fila 6 del Excel)
    DATA_START_ROW = 7
    
    FILAS_GRISES = hojas(ws)

    for row in range(DATA_START_ROW, max_row + 1):
        for col in range(3, max_col + 5):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.font = Contenido

            if col == 3:
                cell.alignment = Alignment(horizontal='left', vertical='center')

            if row in FILAS_GRISES:
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.fill = SECTION_FILL

                if col in range(4, max_col + 5):
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.number_format = FORMATO_NUMERICO_FINANCIERO
            
            if col in range(4, max_col + 5):
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.number_format = FORMATO_NUMERICO_FINANCIERO
            
            if row == 9:
                cell = ws.cell(row=DATA_START_ROW, column=col) 
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    limpiar_rango_Formato(ws, 'C77:I81')

def FormatoPatrimonio(wb, nroHoja, nombre):
    ws = nroHoja
    nombreEmpresa = nombre

    max_row = ws.max_row
    max_col = ws.max_column

    ReordenarTabla(ws, max_col)
    
    # Ajustar el ancho de las columnas
    ws.column_dimensions[get_column_letter(4)].width = 14.5
    for i in range(5, max_col + 5):
        ws.column_dimensions[get_column_letter(i)].width = 10

    # Título Principal (Fila 1, Columna A)
    ws['A1'].font = fuente_titulo
    ws['A3'].value = "Periodo: Anual"
    ws['A3'].font = negrita
    ws['A4'].value = f"Empresa: {nombreEmpresa}"
    ws['A4'].font = negrita
    ws['A5'].value = "Tipo: Individual"
    ws['A5'].font = negrita

    # Ajustamos la fila de los encabezados (ej. Fila 6 del Excel)
    DATA_START_ROW = 7
    
    FILAS_GRISES = hojas(ws)

    for row in range(DATA_START_ROW, max_row + 97):
        for col in range(3, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.font = Contenido

            if col == 3:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

            if row in FILAS_GRISES:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.fill = SECTION_FILL

                if col in range(4, max_col + 1):
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.number_format = FORMATO_NUMERICO_FINANCIERO
            
            if col in range(4, max_col + 1):
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.number_format = FORMATO_NUMERICO_FINANCIERO
            
            if row == 9:
                cell = ws.cell(row=DATA_START_ROW, column=col) 
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws.column_dimensions[get_column_letter(1)].width = 5
    ws.column_dimensions[get_column_letter(2)].width = 5
    ws.column_dimensions[get_column_letter(3)].width = 10

    """
    ws.merge_cells('B8:B31')    # COMBINAR Y CENTRAR - AÑO 2024
    celda2024 = ws['B8']
    celda2024.value = "2024"
    encabezadosFechasVerticales(celda2024)

    ws.merge_cells('B32:B55')    # COMBINAR Y CENTRAR - AÑO 2023
    celda2023 = ws['B32']
    celda2023.value = "2023"
    encabezadosFechasVerticales(celda2023)

    ws.merge_cells('B56:B79')    # COMBINAR Y CENTRAR - AÑO 2023
    celda2022 = ws['B56']
    celda2022.value = "2022"
    encabezadosFechasVerticales(celda2022)

    ws.merge_cells('B80:B103')    # COMBINAR Y CENTRAR - AÑO 2021
    celda2021 = ws['B80']
    celda2021.value = "2021"
    encabezadosFechasVerticales(celda2021)

    ws.merge_cells('B104:B127')    # COMBINAR Y CENTRAR - AÑO 2020
    celda2020 = ws['B104']
    celda2020.value = "2020"
    encabezadosFechasVerticales(celda2020)

    ws.merge_cells('B128:B151')    # COMBINAR Y CENTRAR - AÑO 2019
    celda2019 = ws['B128']
    celda2019.value = "2019"
    encabezadosFechasVerticales(celda2019)
    """

def limpiar_rango_Formato(ws, rango_excel):
    min_col, min_row, max_col, max_row = range_boundaries(rango_excel)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.value = None

def encabezadosFechasVerticales(celda):
    celda.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    celda.font = Font(size=16, bold=True)
    celda.alignment = Alignment(
        horizontal='center',
        vertical='center',
        text_rotation=90  # Rota el texto 90 grados (hacia arriba)
    )
    celda.border = THIN_BORDER

def ReordenarTabla(ws, max_col):    
    # 1. Definir los límites
    # Usaremos A y la última columna (ej. E)
    MAX_COL_LETTER = get_column_letter(max_col + 1)
    MAX_ROW = ws.max_row

    RANGO_SUPERIOR = f"C8:{MAX_COL_LETTER}31"
    ws.move_range(
        RANGO_SUPERIOR,
        rows=+48,
        cols=0
    )

    RANGO_INFERIOR = f"C32:{MAX_COL_LETTER}{MAX_ROW}"
    ws.move_range(
        RANGO_INFERIOR, 
        rows=-24,
        cols=0
    )

    RANGO_SUPERIOR2 = f"C56:{MAX_COL_LETTER}79"
    ws.move_range(
        RANGO_SUPERIOR2,
        rows=-24,
        cols=0
    )

def union_archivos(path_xlsx_origen, path_xlsx_destino, columna):
    # 1. CARGAR AMBOS WORKBOOKS
    wb_origen = load_workbook(path_xlsx_origen, data_only=True)
    wb_destino = load_workbook(path_xlsx_destino)
    
    # 2. DEFINIR HOJAS DE ORIGEN Y DESTINO
    ws_Origen_hoja1 = wb_origen['Hoja1']
    ws_Destino_hoja1 = wb_destino['Hoja1']
    ws_Origen_hoja2 = wb_origen['Hoja2']
    ws_Destino_hoja2 = wb_destino['Hoja2']
    ws_Origen_hoja3 = wb_origen['Hoja3']
    ws_Destino_hoja3 = wb_destino['Hoja3']
    ws_Origen_hoja4 = wb_origen['Hoja4']
    ws_Destino_hoja4 = wb_destino['Hoja4']
    
    # 3. DEFINIR RANGOS COMPLETOS HASTA FILA 83
    # Para hojas 1, 2 y 4 - rango completo hasta fila 83
    rango_completo = 'D7:H104'
    
    # Para hoja 3 (Patrimonio) - rango específico
    rango_patrimonio = 'C8:AA55'  # Ajusta este rango según necesites
    
    fila_destino_inicial = 7
    fila_destino_patrimonio_2023 = 56  # Ajusta según tu estructura
    fila_destino_patrimonio_2022 = 104  # Ajusta según tu estructura

    # 4. EJECUTAR COPIA PARA TODOS LOS AÑOS
    print(f"Copiando datos a columna {columna}...")
    
    # HOJA 1 - SITUACIÓN FINANCIERA
    copiar_celdas_completo(
        ws_Origen_hoja1,
        ws_Destino_hoja1,
        rango_completo,
        fila_destino_inicial,
        columna
    )
    
    # HOJA 2 - RESULTADOS
    copiar_celdas_completo(
        ws_Origen_hoja2,
        ws_Destino_hoja2,
        rango_completo,
        fila_destino_inicial,
        columna
    )
    
    # HOJA 3 - PATRIMONIO (reactivar esta sección)
    # Determinar posición destino basado en la columna
    if columna == 5:  # 2024
        fila_destino_patrimonio = 8
        columna_destino_patrimonio = 4
    elif columna == 6:  # 2023
        fila_destino_patrimonio = 32
        columna_destino_patrimonio = 4
    elif columna == 7:  # 2022
        fila_destino_patrimonio = 56
        columna_destino_patrimonio = 4
    elif columna == 8:  # 2021
        fila_destino_patrimonio = 80
        columna_destino_patrimonio = 4
    elif columna == 9:  # 2020
        fila_destino_patrimonio = 104
        columna_destino_patrimonio = 4
    else:  # 2019 u otros
        fila_destino_patrimonio = 128
        columna_destino_patrimonio = 4
    
    copiar_celdas_completo(
        ws_Origen_hoja3,
        ws_Destino_hoja3,
        rango_patrimonio,
        fila_destino_patrimonio,
        columna_destino_patrimonio
    )
    
    # HOJA 4 - FLUJO DE EFECTIVO
    copiar_celdas_completo(
        ws_Origen_hoja4,
        ws_Destino_hoja4,
        rango_completo,
        fila_destino_inicial,
        columna
    )
    
    # 5. GUARDAR EL ARCHIVO DESTINO
    wb_destino.save(path_xlsx_destino)
    print(f"Datos copiados exitosamente a columna {columna}")

def copiar_celdas_completo(ws_origen, ws_destino, rango_origen: str, fila_inicio_destino: int, columna_inicio_destino: int):
    """Versión mejorada que asegura copia completa hasta la fila 83"""
    
    # Obtener las coordenadas del rango de origen
    try:
        min_col, min_row, max_col, max_row = range_boundaries(rango_origen) 
    except ValueError:
        print(f"Error: Rango '{rango_origen}' no es válido.")
        return

    # Verificar que el rango de origen tenga datos hasta la fila 83
    fila_maxima_origen = ws_origen.max_row
    if max_row > fila_maxima_origen:
        print(f"Advertencia: Rango solicitado hasta fila {max_row} pero origen solo tiene hasta {fila_maxima_origen}")
        max_row = fila_maxima_origen

    fila_destino = fila_inicio_destino
    celdas_copiadas = 0
    
    # Iterar sobre las filas y columnas del rango de origen
    for row_num in range(min_row, max_row + 1):
        col_destino = columna_inicio_destino
        
        for col_num in range(min_col, max_col + 1):
            cell_origen = ws_origen.cell(row=row_num, column=col_num)
            cell_destino = ws_destino.cell(row=fila_destino, column=col_destino)
            
            # Copiar valor
            cell_destino.value = cell_origen.value 
            
            # Copiar formato numérico para celdas numéricas
            if isinstance(cell_origen.value, (int, float)):
                cell_destino.number_format = FORMATO_NUMERICO_FINANCIERO
            
            celdas_copiadas += 1
            col_destino += 1
            
        fila_destino += 1
    
    print(f"Copiadas {celdas_copiadas} celdas desde fila {min_row} a {max_row}")

# Función adicional para diagnóstico
def verificar_rangos(path_archivo):
    """Función para diagnosticar los rangos reales de datos"""
    wb = load_workbook(path_archivo, data_only=True)
    
    for sheet_name in ['Hoja1', 'Hoja2', 'Hoja3', 'Hoja4']:
        ws = wb[sheet_name]
        print(f"\n{sheet_name}:")
        print(f"  - Máxima fila con datos: {ws.max_row}")
        print(f"  - Máxima columna con datos: {ws.max_column}")
        
        # Verificar específicamente las filas alrededor de 83
        for row in [80, 81, 82, 83, 84]:
            has_data = any(ws.cell(row=row, column=col).value is not None 
                          for col in range(1, ws.max_column + 1))
            if has_data:
                print(f"  - Fila {row}: TIENE DATOS")
            else:
                print(f"  - Fila {row}: vacía")

def extraer_ratios(ws, fila_inicio, fila_fin, col_inicio, col_fin):
    """
    Convierte la tabla de ratios en un diccionario, filtrando encabezados de sección
    """
    datos_ratios = {}

    # Encabezados (años)
    anios = []
    for col in range(col_inicio, col_fin+1):
        val = ws.cell(row=fila_inicio, column=col).value
        if isinstance(val, (int, float)):
            val = str(int(val))
        else:
            val = str(val).strip() if val else ""
        anios.append(val)

    # Lista de encabezados de sección a excluir
    encabezados_excluir = [
        "RATIOS DE LIQUIDEZ", 
        "RATIOS DE GESTIÓN", 
        "RATIOS DE ENDEUDAMIENTO", 
        "RATIOS DE RENTABILIDAD",
        "RATIOS DE LIQUIDEZ Y SOLVENCIA",
        "RATIOS DE GESTIÓN O ACTIVIDAD", 
        "RATIOS DE ENDEUDAMIENTO O APALANCAMIENTO",
        "RATIOS DE RENTABILIDAD O RENDIMIENTO"
    ]

    # Iterar cada fila de ratios
    for fila in range(fila_inicio+1, fila_fin+1):
        nombre_ratio = ws.cell(row=fila, column=col_inicio-1).value
        
        # Filtrar filas vacías y encabezados de sección
        if not nombre_ratio:
            continue
            
        nombre_ratio = str(nombre_ratio).strip()
        
        # Excluir encabezados de sección (en mayúsculas generalmente)
        if (nombre_ratio.upper() in [e.upper() for e in encabezados_excluir] or 
            nombre_ratio.startswith('RATIOS DE')):
            continue

        datos_ratios[nombre_ratio] = {}

        # Extraer valores, manejando None
        for idx, col in enumerate(range(col_inicio, col_fin+1)):
            valor = ws.cell(row=fila, column=col).value
            
            # Convertir a float si es numérico, mantener None si es texto vacío
            if isinstance(valor, (int, float)):
                datos_ratios[nombre_ratio][anios[idx]] = float(valor)
            elif valor and str(valor).strip().replace('.', '').isdigit():
                datos_ratios[nombre_ratio][anios[idx]] = float(valor)
            else:
                datos_ratios[nombre_ratio][anios[idx]] = 0.0  # Default a 0 en lugar de None

    return datos_ratios

def copiar_celdas(ws_origen, ws_destino, rango_origen: str, fila_inicio_destino: int, columna_inicio_destino: int):    
    # Obtener las coordenadas del rango de origen
    try:
        min_col, min_row, max_col, max_row = range_boundaries(rango_origen) 
    except ValueError:
        print(f"Error: Rango '{rango_origen}' no es válido.")
        return

    fila_destino = fila_inicio_destino
    
    # Iterar sobre las filas y columnas del rango de origen
    for row in ws_origen.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        col_destino = columna_inicio_destino
        
        for cell_origen in row:
            # Obtener la celda de destino
            cell_destino = ws_destino.cell(row=fila_destino, column=col_destino)
            
            # 1. Copiar Valor: Solo si la celda tiene un valor (ignora celdas combinadas secundarias)
            if cell_origen.value is not None:
                cell_destino.value = cell_origen.value 
                
                # 2. Copiar Formato Numérico
                if cell_origen.number_format:
                    cell_destino.number_format = FORMATO_NUMERICO_FINANCIERO
            else:
                # Opcional: Asegurar que la celda destino también esté vacía si el origen lo está
                cell_destino.value = None

            col_destino += 1
        fila_destino += 1


def to_number(value):
    """Convierte un valor a float si es posible, maneja porcentajes, comas, textos y paréntesis para negativos"""
    if value is None:
        return None
        
    if isinstance(value, (int, float)):
        return float(value)
    
    try:
        if isinstance(value, str):
            value = value.strip()
            if value == '' or value == '-':
                return None
            
            # Manejar paréntesis para negativos - MÁS ROBUSTO
            # Caso 1: (123.45) -> -123.45
            # Caso 2: (1,234.56) -> -1234.56
            # Caso 3: (123) -> -123
            if '(' in value and ')' in value:
                # Extraer el contenido dentro de paréntesis
                start = value.find('(')
                end = value.find(')')
                if start < end:  # Verificar que los paréntesis estén en orden correcto
                    content = value[start+1:end].strip()
                    
                    # Limpiar el contenido (quitar comas, espacios, símbolos de moneda)
                    content = content.replace(',', '').replace(' ', '')
                    content = content.replace('$', '').replace('€', '').replace('£', '').replace('¥', '')
                    
                    # Si el contenido es un número válido, convertirlo a negativo
                    if content and content != '-':
                        try:
                            return -float(content)
                        except ValueError:
                            # Si falla, continuar con el procesamiento normal
                            value = value.replace('(', '').replace(')', '')
                            value = '-' + value.strip()
            
            # Manejar el formato negativo con paréntesis ya procesado
            # o continuar con el procesamiento normal si los paréntesis no eran válidos
            
            # Manejar porcentajes
            if '%' in value:
                value = value.replace('%', '').strip()
                result = to_number(value)
                return result / 100 if result is not None else None
            
            # Manejar múltiples formatos de paréntesis
            # Caso: - (123.45) algunos formatos usan espacio después del signo negativo
            if value.startswith('- (') and value.endswith(')'):
                value = '-' + value[3:-1]
            elif value.startswith('-(') and value.endswith(')'):
                value = '-' + value[2:-1]
            
            # Manejar comas decimales
            if ',' in value and '.' in value:
                # Caso: 1,234.56 (coma como separador de miles, punto decimal)
                value = value.replace(',', '')
            elif ',' in value:
                # Caso: 1234,56 (coma como decimal)
                # Verificar si la coma es separador decimal o de miles
                parts = value.split(',')
                if len(parts) == 2 and len(parts[1]) <= 2:
                    # Probablemente coma decimal (ej: 1234,56)
                    value = value.replace(',', '.')
                else:
                    # Probablemente coma como separador de miles (ej: 1,234)
                    value = value.replace(',', '')
            
            # Limpiar caracteres no numéricos
            value = value.replace(' ', '')
            value = value.replace('$', '').replace('€', '').replace('£', '').replace('¥', '')
            
            # Manejar signo negativo al inicio
            if value.startswith('--'):  # Caso de doble negativo
                value = value.replace('--', '-')
            
            if value == '' or value == '-':
                return None
            
            # Convertir a float
            return float(value)
            
    except (ValueError, TypeError) as e:
        print(f"Error convirtiendo valor '{value}': {e}")
        return None
    
    return None

# Función adicional para debugging
def debug_to_number(value):
    """Versión de debug para ver qué está pasando con la conversión"""
    print(f"Convirtiendo: '{value}' (tipo: {type(value)})")
    result = to_number(value)
    print(f"Resultado: {result}")
    return result

# Ejemplos de prueba
def test_to_number():
    """Prueba la función to_number con diferentes formatos"""
    test_cases = [
        "(123.45)",           # -> -123.45
        "(1,234.56)",         # -> -1234.56
        "(123)",              # -> -123
        " (456.78) ",         # -> -456.78
        "- (789.12)",         # -> -789.12
        "-(345.67)",          # -> -345.67
        "1,234.56",           # -> 1234.56
        "1234,56",            # -> 1234.56
        "-123.45",            # -> -123.45
        "123.45",             # -> 123.45
        "$1,234.56",          # -> 1234.56
        "(€1,234.56)",        # -> -1234.56
        " - 123.45 ",         # -> -123.45
        "123.45%",            # -> 1.2345
        "(123.45%)",          # -> -1.2345
        "",                   # -> None
        "-",                  # -> None
        "ABC",                # -> None
        None,                 # -> None
    ]
    
    print("=== PRUEBAS DE to_number ===")
    for test in test_cases:
        result = to_number(test)
        print(f"'{test}' -> {result}")

# Mejora adicional: función específica para valores financieros
def to_financial_number(value):
    """Versión especializada para valores financieros con paréntesis"""
    if value is None:
        return None
    
    if isinstance(value, (int, float)):
        return float(value)
    
    if isinstance(value, str):
        value = value.strip()
        
        # Caso especial: paréntesis que representan negativo
        if value.startswith('(') and value.endswith(')'):
            # Extraer el número dentro de paréntesis y hacerlo negativo
            inner_value = value[1:-1].strip()
            # Limpiar comas y espacios
            inner_value = inner_value.replace(',', '').replace(' ', '')
            inner_value = inner_value.replace('$', '').replace('€', '').replace('£', '').replace('¥', '')
            
            try:
                return -float(inner_value)
            except (ValueError, TypeError):
                # Si falla, usar el método normal
                pass
        
        # Para otros casos, usar la función principal
        return to_number(value)
    
    return None

def encontrar_total_activo_por_columna(ws, columna):
    """Encuentra el TOTAL DEL ACTIVO en una columna específica"""
    for row in range(1, 120):
        cell_value = ws[f'C{row}'].value
        if cell_value and "TOTAL DEL ACTIVO" in str(cell_value).upper():
            total_value = to_number(ws[f'{columna}{row}'].value)
            if total_value not in (None, 0):
                return total_value, f'{columna}{row}'
    return None, None

def AnalisisVerticalMultiAnio(ws):
    """Aplica análisis vertical para todos los años (2024-2020)"""
    print(f"Aplicando análisis vertical multi-año en {ws.title}")
    
    columnas_anios = {'D': '2024', 'E': '2023', 'F': '2022', 'G': '2021', 'H': '2020'}
    columnas_destino = {'D': 'J', 'E': 'K', 'F': 'L', 'G': 'M', 'H': 'N'}
    
    cambios_totales = 0
    
    for col_origen, anio in columnas_anios.items():
        print(f"Procesando año {anio}...")
        
        total_activo, ubicacion_total = encontrar_total_activo_por_columna(ws, col_origen)
        col_destino = columnas_destino[col_origen]
        
        if total_activo not in (None, 0):
            cambios_anio = 0
            
            for row in range(8, 105):
                try:
                    # Obtener la fila del total activo
                    fila_total = int(ubicacion_total.replace(col_origen, '')) if ubicacion_total else -1
                    
                    # Si es la fila del TOTAL DEL ACTIVO, poner 100%
                    if row == fila_total:
                        ws[f'{col_destino}{row}'].value = 1.0  # 100%
                        ws[f'{col_destino}{row}'].number_format = "0.00%"
                        cambios_anio += 1
                        continue
                    
                    valor = to_number(ws[f'{col_origen}{row}'].value)
                    
                    if valor == 0:
                        ws[f'{col_destino}{row}'].value = 0
                        ws[f'{col_destino}{row}'].number_format = "0.00%"
                        cambios_anio += 1
                    elif valor is not None:
                        resultado = valor / total_activo
                        ws[f'{col_destino}{row}'].value = resultado
                        ws[f'{col_destino}{row}'].number_format = "0.00%"
                        cambios_anio += 1
                        
                except Exception as e:
                    continue
            
            print(f"✓ {anio}: {cambios_anio} cálculos")
            cambios_totales += cambios_anio
    
    print(f"Análisis vertical completado: {cambios_totales} cálculos")
    return cambios_totales

def AnalisisHorizontalMultiAnio(ws):
    """Aplica análisis horizontal para todas las comparaciones entre años"""
    print(f"Aplicando análisis horizontal multi-año en {ws.title}")
    
    comparaciones = {
        ('D', 'E'): 'P',  # 2024 vs 2023
        ('E', 'F'): 'Q',  # 2023 vs 2022  
        ('F', 'G'): 'R',  # 2022 vs 2021
        ('G', 'H'): 'S'   # 2021 vs 2020
    }
    
    cambios_totales = 0
    
    for (col_actual, col_anterior), col_destino in comparaciones.items():
        cambios_comparacion = 0
        
        for row in range(8, 105):
            try:
                valor_actual = to_number(ws[f'{col_actual}{row}'].value)
                valor_anterior = to_number(ws[f'{col_anterior}{row}'].value)
                
                # Ambos cero
                if valor_actual == 0 and valor_anterior == 0:
                    ws[f'{col_destino}{row}'].value = 0
                    ws[f'{col_destino}{row}'].number_format = "0.00%"
                    cambios_comparacion += 1
                # Año anterior cero, actual tiene valor
                elif valor_anterior == 0 and valor_actual is not None and valor_actual != 0:
                    ws[f'{col_destino}{row}'].value = None
                # Año actual cero, anterior tenía valor
                elif valor_actual == 0 and valor_anterior is not None and valor_anterior != 0:
                    ws[f'{col_destino}{row}'].value = -1
                    ws[f'{col_destino}{row}'].number_format = "0.00%"
                    cambios_comparacion += 1
                # Cálculo normal
                elif (valor_actual is not None and valor_anterior is not None and valor_anterior != 0):
                    resultado = (valor_actual - valor_anterior) / valor_anterior
                    ws[f'{col_destino}{row}'].value = resultado
                    ws[f'{col_destino}{row}'].number_format = "0.00%"
                    cambios_comparacion += 1
                    
            except Exception:
                continue
        
        print(f"✓ {col_actual}vs{col_anterior}: {cambios_comparacion} cálculos")
        cambios_totales += cambios_comparacion
    
    print(f"Análisis horizontal completado: {cambios_totales} cálculos")
    return cambios_totales

def AnalisisVerticalFlujoEfectivo(ws):
    """Aplica análisis vertical para Estado de Flujo de Efectivo"""
    print(f"Aplicando análisis vertical a Flujo de Efectivo en {ws.title}")
    
    columnas_anios = {
        'D': '2024',
        'E': '2023', 
        'F': '2022',
        'G': '2021',
        'H': '2020'
    }
    
    columnas_destino = {
        'D': 'J', 'E': 'K', 'F': 'L', 'G': 'M', 'H': 'N'
    }
    
    cambios_totales = 0
    
    for col_origen, anio in columnas_anios.items():
        print(f"Procesando año {anio} para Flujo de Efectivo...")
        
        total_base, ubicacion_total = encontrar_flujo_operacion_por_columna(ws, col_origen)
        col_destino = columnas_destino[col_origen]
        
        if total_base not in (None, 0):
            cambios_anio = 0
            
            print(f"FLUJO OPERACIÓN {anio}: {total_base:,.0f}")
            
            for row in range(8, 77):
                try:
                    # Obtener la fila del flujo de operación
                    fila_total = int(ubicacion_total.replace(col_origen, '')) if ubicacion_total else -1
                    
                    # Si es la fila del FLUJO OPERACIÓN, poner 100% (o -100% si es negativo)
                    if row == fila_total:
                        if total_base >= 0:
                            ws[f'{col_destino}{row}'].value = 1.0  # 100%
                        else:
                            ws[f'{col_destino}{row}'].value = -1.0  # -100%
                        ws[f'{col_destino}{row}'].number_format = "0.00%"
                        cambios_anio += 1
                        continue
                    
                    valor = to_number(ws[f'{col_origen}{row}'].value)
                    
                    if valor == 0:
                        ws[f'{col_destino}{row}'].value = 0
                        ws[f'{col_destino}{row}'].number_format = "0.00%"
                        cambios_anio += 1
                    elif valor is not None:
                        # Usar valor absoluto del flujo de operación como base
                        resultado = valor / abs(total_base)
                        ws[f'{col_destino}{row}'].value = resultado
                        ws[f'{col_destino}{row}'].number_format = "0.00%"
                        cambios_anio += 1
                        
                except Exception as e:
                    print(f"Error en fila {row} para {anio}: {e}")
            
            print(f"✓ Análisis vertical Flujo Efectivo {anio}: {cambios_anio} cálculos")
            cambios_totales += cambios_anio
        else:
            print(f"⚠ No se encontró FLUJO OPERACIÓN válido para {anio}")
    
    print(f"Análisis vertical Flujo de Efectivo completado: {cambios_totales} cálculos")
    return cambios_totales

def get_anio_from_columna(columna):
    """Convierte letra de columna a año"""
    mapeo = {'D': '2024', 'E': '2023', 'F': '2022', 'G': '2021', 'H': '2020'}
    return mapeo.get(columna, 'Desconocido')

def AnalisisVerticalResultados(ws):
    """Aplica análisis vertical para Estado de Resultados usando TOTAL INGRESOS POR INTERESES como base"""
    print(f"Aplicando análisis vertical a Estado de Resultados en {ws.title}")
    
    columnas_anios = {
        'D': '2024', 'E': '2023', 'F': '2022', 'G': '2021', 'H': '2020'
    }
    
    columnas_destino = {
        'D': 'J', 'E': 'K', 'F': 'L', 'G': 'M', 'H': 'N'
    }
    
    cambios_totales = 0
    
    for col_origen, anio in columnas_anios.items():
        print(f"Procesando año {anio} para Estado de Resultados...")
        
        total_base, ubicacion_total = encontrar_total_ingresos_intereses_por_columna(ws, col_origen)
        col_destino = columnas_destino[col_origen]
        
        if total_base not in (None, 0):
            cambios_anio = 0
            
            print(f"TOTAL INGRESOS POR INTERESES {anio}: {total_base:,.0f}")
            
            for row in range(8, 50):
                try:
                    # Obtener la fila del total base
                    fila_total = int(ubicacion_total.replace(col_origen, '')) if ubicacion_total else -1
                    
                    # Si es la fila del TOTAL INGRESOS POR INTERESES, poner 100%
                    if row == fila_total:
                        ws[f'{col_destino}{row}'].value = 1.0  # 100%
                        ws[f'{col_destino}{row}'].number_format = "0.00%"
                        cambios_anio += 1
                        continue
                    
                    valor = to_number(ws[f'{col_origen}{row}'].value)
                    
                    if valor == 0:
                        ws[f'{col_destino}{row}'].value = 0
                        ws[f'{col_destino}{row}'].number_format = "0.00%"
                        cambios_anio += 1
                    elif valor is not None:
                        resultado = valor / abs(total_base)
                        ws[f'{col_destino}{row}'].value = resultado
                        ws[f'{col_destino}{row}'].number_format = "0.00%"
                        cambios_anio += 1
                        
                except Exception as e:
                    print(f"Error en fila {row} para {anio}: {e}")
            
            print(f"✓ Análisis vertical Resultados {anio}: {cambios_anio} cálculos")
            cambios_totales += cambios_anio
        else:
            print(f"⚠ No se encontró TOTAL INGRESOS POR INTERESES válido para {anio}")
    
    print(f"Análisis vertical Estado de Resultados completado: {cambios_totales} cálculos")
    return cambios_totales

def encontrar_total_ingresos_intereses_por_columna(ws, columna):
    """Encuentra el TOTAL INGRESOS POR INTERESES en una columna específica"""
    nombres_ingresos_intereses = [
        "TOTAL INGRESOS POR INTERESES",
        "TOTAL INGRESOS FINANCIEROS",
        "INGRESOS POR INTERESES TOTALES"
    ]
    
    for row in range(1, 50):
        cell_value = ws[f'C{row}'].value
        if cell_value:
            cell_upper = str(cell_value).upper()
            for nombre in nombres_ingresos_intereses:
                if nombre in cell_upper:
                    total_value = to_number(ws[f'{columna}{row}'].value)
                    if total_value not in (None, 0):
                        print(f"✓ Total Ingresos por Intereses encontrado en C{row}: {total_value:,.0f}")
                        return total_value, f'{columna}{row}'
    
    # Si no se encuentra, buscar INGRESOS POR INTERESES como alternativa
    print("Buscando INGRESOS POR INTERESES como alternativa...")
    for row in range(1, 50):
        cell_value = ws[f'C{row}'].value
        if cell_value and "INGRESOS POR INTERESES" in str(cell_value).upper():
            # Sumar los ingresos por intereses individuales
            total_value = 0
            current_row = row + 1
            while current_row < 50:
                next_cell = ws[f'C{current_row}'].value
                if next_cell and "TOTAL INGRESOS POR INTERESES" in str(next_cell).upper():
                    total_value = to_number(ws[f'{columna}{current_row}'].value)
                    if total_value not in (None, 0):
                        print(f"✓ Total Ingresos por Intereses (suma) encontrado en C{current_row}: {total_value:,.0f}")
                        return total_value, f'{columna}{current_row}'
                current_row += 1
    
    print("✗ No se pudo encontrar base para análisis vertical de estado de resultados")
    return None, None

def AnalisisHorizontalResultados(ws):
    """Aplica análisis horizontal para Estado de Resultados"""
    print(f"Aplicando análisis horizontal a Estado de Resultados en {ws.title}")
    
    comparaciones = {
        ('D', 'E'): 'P',  # 2024 vs 2023
        ('E', 'F'): 'Q',  # 2023 vs 2022  
        ('F', 'G'): 'R',  # 2022 vs 2021
        ('G', 'H'): 'S'   # 2021 vs 2020
    }
    
    cambios_totales = 0
    
    for (col_actual, col_anterior), col_destino in comparaciones.items():
        cambios_comparacion = 0
        anio_actual = get_anio_from_columna(col_actual)
        anio_anterior = get_anio_from_columna(col_anterior)
        
        print(f"Análisis horizontal {anio_actual} vs {anio_anterior} para Resultados...")
        
        for row in range(8, 50):  # Rango del estado de resultados
            try:
                valor_actual = to_number(ws[f'{col_actual}{row}'].value)
                valor_anterior = to_number(ws[f'{col_anterior}{row}'].value)
                
                # Ambos cero
                if valor_actual == 0 and valor_anterior == 0:
                    ws[f'{col_destino}{row}'].value = 0
                    ws[f'{col_destino}{row}'].number_format = "0.00%"
                    cambios_comparacion += 1
                
                # Año anterior cero, actual tiene valor
                elif valor_anterior == 0 and valor_actual is not None and valor_actual != 0:
                    ws[f'{col_destino}{row}'].value = None  # No calculable
                
                # Año actual cero, anterior tenía valor
                elif valor_actual == 0 and valor_anterior is not None and valor_anterior != 0:
                    ws[f'{col_destino}{row}'].value = -1  # -100%
                    ws[f'{col_destino}{row}'].number_format = "0.00%"
                    cambios_comparacion += 1
                
                # Cálculo normal
                elif (valor_actual is not None and valor_anterior is not None and valor_anterior != 0):
                    resultado = (valor_actual - valor_anterior) / valor_anterior
                    
                    # Limitar valores extremos para evitar ####
                    if resultado > 100:  # Más de 10000%
                        resultado = 100
                    elif resultado < -1:  # Menos de -100%
                        resultado = -1
                    
                    ws[f'{col_destino}{row}'].value = resultado
                    ws[f'{col_destino}{row}'].number_format = "0.00%"
                    cambios_comparacion += 1
                    
            except Exception as e:
                print(f"Error en fila {row} para {anio_actual} vs {anio_anterior}: {e}")
        
        print(f"✓ {anio_actual} vs {anio_anterior}: {cambios_comparacion} cálculos")
        cambios_totales += cambios_comparacion
    
    print(f"Análisis horizontal Estado de Resultados completado: {cambios_totales} cálculos")
    return cambios_totales

def encontrar_flujo_operacion_por_columna(ws, columna):
    """Encuentra el FLUJO NETO DE ACTIVIDADES DE OPERACIÓN en una columna específica"""
    # Buscar por diferentes nombres posibles
    nombres_flujo_operacion = [
        "FLUJOS DE EFECTIVO NETO DE ACTIVIDADES DE OPERACIÓN",
        "FLUJO NETO DE ACTIVIDADES DE OPERACIÓN", 
        "FLUJOS NETOS DE OPERACIÓN",
        "FLUJO DE EFECTIVO NETO DE OPERACIÓN"
    ]
    
    for row in range(1, 100):
        cell_value = ws[f'C{row}'].value
        if cell_value:
            cell_upper = str(cell_value).upper()
            for nombre in nombres_flujo_operacion:
                if nombre in cell_upper:
                    total_value = to_number(ws[f'{columna}{row}'].value)
                    if total_value not in (None, 0):
                        print(f"✓ Flujo Operación encontrado en C{row}: {total_value:,.0f}")
                        return total_value, f'{columna}{row}'
    
    # Si no se encuentra, buscar el resultado neto como alternativa
    print("Buscando Resultado Neto como alternativa...")
    for row in range(1, 100):
        cell_value = ws[f'C{row}'].value
        if cell_value and "RESULTADO NETO DEL EJERCICIO" in str(cell_value).upper():
            total_value = to_number(ws[f'{columna}{row}'].value)
            if total_value not in (None, 0):
                print(f"✓ Usando Resultado Neto en C{row}: {total_value:,.0f}")
                return total_value, f'{columna}{row}'
    
    print("✗ No se pudo encontrar base para análisis vertical de flujo de efectivo")
    return None, None

def AnalisisHorizontalFlujoEfectivo(ws):
    """Aplica análisis horizontal para Estado de Flujo de Efectivo"""
    print(f"Aplicando análisis horizontal a Flujo de Efectivo en {ws.title}")
    
    comparaciones = {
        ('D', 'E'): 'P',  # 2024 vs 2023
        ('E', 'F'): 'Q',  # 2023 vs 2022  
        ('F', 'G'): 'R',  # 2022 vs 2021
        ('G', 'H'): 'S'   # 2021 vs 2020
    }
    
    cambios_totales = 0
    
    for (col_actual, col_anterior), col_destino in comparaciones.items():
        cambios_comparacion = 0
        anio_actual = get_anio_from_columna(col_actual)
        anio_anterior = get_anio_from_columna(col_anterior)
        
        print(f"Análisis horizontal {anio_actual} vs {anio_anterior} para Flujo Efectivo...")
        
        for row in range(8, 77):  # Rango del flujo de efectivo
            try:
                valor_actual = to_number(ws[f'{col_actual}{row}'].value)
                valor_anterior = to_number(ws[f'{col_anterior}{row}'].value)
                
                # Ambos cero
                if valor_actual == 0 and valor_anterior == 0:
                    ws[f'{col_destino}{row}'].value = 0
                    ws[f'{col_destino}{row}'].number_format = "0.00%"
                    cambios_comparacion += 1
                
                # Año anterior cero, actual tiene valor
                elif valor_anterior == 0 and valor_actual is not None and valor_actual != 0:
                    ws[f'{col_destino}{row}'].value = None  # No calculable
                
                # Año actual cero, anterior tenía valor
                elif valor_actual == 0 and valor_anterior is not None and valor_anterior != 0:
                    ws[f'{col_destino}{row}'].value = -1  # -100%
                    ws[f'{col_destino}{row}'].number_format = "0.00%"
                    cambios_comparacion += 1
                
                # Cálculo normal
                elif (valor_actual is not None and valor_anterior is not None and valor_anterior != 0):
                    resultado = (valor_actual - valor_anterior) / valor_anterior
                    
                    # Limitar valores extremos para evitar ####
                    if resultado > 100:  # Más de 10000%
                        resultado = 100
                    elif resultado < -1:  # Menos de -100%
                        resultado = -1
                    
                    ws[f'{col_destino}{row}'].value = resultado
                    ws[f'{col_destino}{row}'].number_format = "0.00%"
                    cambios_comparacion += 1
                    
            except Exception as e:
                print(f"Error en fila {row} para {anio_actual} vs {anio_anterior}: {e}")
        
        print(f"✓ {anio_actual} vs {anio_anterior}: {cambios_comparacion} cálculos")
        cambios_totales += cambios_comparacion
    
    print(f"Análisis horizontal Flujo de Efectivo completado: {cambios_totales} cálculos")
    return cambios_totales


def analisis_VH(path_xlsx):
    try:
        wb = load_workbook(path_xlsx)
        
        print("=== INICIANDO ANÁLISIS VH MULTI-AÑO ===")
        print(f"Hojas disponibles: {wb.sheetnames}")
        
        # Verificar que las hojas de datos existen
        hojas_datos = ['Hoja1', 'Hoja2', 'Hoja4']
        for hoja in hojas_datos:
            if hoja not in wb.sheetnames:
                print(f"ERROR: No se encuentra {hoja} en el archivo")
                return False
        
        balance = wb['Hoja1']
        resultados = wb['Hoja2']
        flujos = wb['Hoja4']
        
        # CREAR O MANTENER Hoja5 - NO SE ELIMINA
        if 'Hoja5' not in wb.sheetnames:
            Ratios = wb.copy_worksheet(balance)
            Ratios.title = "Hoja5"
            Ratios['A1'].value = "RATIOS FINANCIEROS"
            print("✓ Hoja5 creada para ratios")
        else:
            Ratios = wb['Hoja5']
            print("✓ Hoja5 ya existe - SE CONSERVA")
        
        # Aplicar formatos
        print("\n1. APLICANDO FORMATOS...")
        FormatoAnalisis1(balance)
        FormatoAnalisis2(resultados)
        FormatoAnalisis3(flujos)
        
        # Aplicar análisis multi-año
        print("\n2. APLICANDO ANÁLISIS MULTI-AÑO...")
        
        print("\n=== BALANCE GENERAL (Hoja1) ===")
        cambios_v_balance = AnalisisVerticalMultiAnio(balance)
        cambios_h_balance = AnalisisHorizontalMultiAnio(balance)
        
        print("\n=== ESTADO DE RESULTADOS (Hoja2) ===")
        cambios_v_resultados = AnalisisVerticalResultados(resultados)  # Nueva función específica
        cambios_h_resultados = AnalisisHorizontalResultados(resultados)  # Nueva función específica
        
        print("\n=== FLUJO DE EFECTIVO (Hoja4) ===")
        cambios_v_flujos = AnalisisVerticalFlujoEfectivo(flujos)
        cambios_h_flujos = AnalisisHorizontalFlujoEfectivo(flujos)
        
        print("\n=== RESUMEN FINAL ===")
        print(f"BALANCE GENERAL (Hoja1):")
        print(f"  Vertical: {cambios_v_balance} cálculos")
        print(f"  Horizontal: {cambios_h_balance} cálculos")
        print(f"ESTADO DE RESULTADOS (Hoja2):")
        print(f"  Vertical: {cambios_v_resultados} cálculos") 
        print(f"  Horizontal: {cambios_h_resultados} cálculos")
        print(f"FLUJO DE EFECTIVO (Hoja4):")
        print(f"  Vertical: {cambios_v_flujos} cálculos")
        print(f"  Horizontal: {cambios_h_flujos} cálculos")
        print(f"HOJA5: CONSERVADA INTACTA PARA RATIOS FUTUROS")
        
        # Guardar cambios
        wb.save(path_xlsx)
        print("\n✓ Archivo guardado correctamente")
        print("✓ Análisis multi-año completado para todos los estados financieros")
        return True
        
    except Exception as e:
        print(f"ERROR en analisis_VH: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def FormatoAnalisis1(ws):
    """Formato para Balance General con colores (igual a FormatoAnalisis2)"""
    if ws.title == 'Hoja5':
        print("✓ Saltando formato para Hoja5")
        return
        
    print(f"Aplicando formato de análisis con colores a {ws.title}")
    
    ws.column_dimensions[get_column_letter(9)].width = 3
    ws.column_dimensions[get_column_letter(15)].width = 3
    
    # Análisis Vertical (5 años)
    ws.merge_cells('J6:N6')
    aplicarBorde(ws, 'J6:N104')
    ws['J6'].value = "Análisis Vertical"
    ws['J6'].fill = ENCABEZADO_NARANJA
    ws['J6'].font = negrita
    ws['J6'].alignment = Alignment(horizontal='center', vertical='center')

    # Análisis Horizontal (4 comparaciones) - CORREGIDO: merge_cells en lugar de merge_cordes
    ws.merge_cells('P6:S6')
    aplicarBorde(ws, 'P6:S104')
    ws['P6'].value = "Análisis Horizontal"
    ws['P6'].fill = ENCABEZADO_NARANJA
    ws['P6'].font = negrita
    ws['P6'].alignment = Alignment(horizontal='center', vertical='center')

    # Copiar encabezados de años
    copiar_celdas(ws, ws, 'D7:H7', 7, 10)
    copiar_celdas(ws, ws, 'D7:G7', 7, 16)

    for row in range(7, 105):
        for col in range(10, 15):
            cell = ws.cell(row=row, column=col)
            if row == 7:
                cell = ws.cell(row=7, column=col) 
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            elif row in [8,9,24,25,43,44,45,46,59,60,72,73,74,82,83]:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.fill = SECTION_FILL
            
            if row in range(8, 105):
                if ((10 <= row <= 25) or (26 <= row <= 44) or (47 <= row <= 59) or 
                    (61 <= row <= 73) or (75 <= row <= 83)):
                    formatoCondicional_EstalaDeColor(ws, row, col, col)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.number_format = FORMATO_PORCENTAJE_DOS_DECIMALES
        
        for col in range(16, 20):
            cell = ws.cell(row=row, column=col)
            if row == 7:
                cell = ws.cell(row=7, column=col) 
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            elif row in [8,9,24,25,43,44,45,46,59,60,72,73,74,82,83]:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.fill = SECTION_FILL
            
            if row in range(8, 105):
                if ((10 <= row <= 25) or (26 <= row <= 44) or (47 <= row <= 59) or 
                    (61 <= row <= 73) or (75 <= row <= 83)):
                    formatoCondicional_EstalaDeColor(ws, row, col, col)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.number_format = FORMATO_PORCENTAJE_DOS_DECIMALES

def FormatoAnalisis2(ws):
    ws.column_dimensions[get_column_letter(9)].width = 3
    ws.column_dimensions[get_column_letter(15)].width = 3
    ws.merge_cells('J6:N6')
    aplicarBorde(ws, 'J6:N32')
    ws['J6'].value = "Análisis Vertical"
    ws['J6'].fill = ENCABEZADO_NARANJA
    ws['J6'].font = negrita
    ws['J6'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('P6:S6')
    aplicarBorde(ws, 'P6:S32')
    ws['P6'].value = "Análisis Horizontal"
    ws['P6'].fill = ENCABEZADO_NARANJA
    ws['P6'].font = negrita
    ws['P6'].alignment = Alignment(horizontal='center', vertical='center')

    copiar_celdas(ws,ws,'D7:H7',7,10)
    copiar_celdas(ws,ws,'D7:G7',7,16)

    for row in range(7, 33):
        for col in range(10, 15):
            cell = ws.cell(row=row, column=col)
            if row == 7:
                cell = ws.cell(row=7, column=col) 
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            elif row in [8,10,16,28,32]:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.fill = SECTION_FILL
            
            if row in range(8, 33):
                if (row == 9 or (11 <= row <= 15) or (17 <= row <= 27) or (29 <= row <= 31)):
                    formatoCondicional_EstalaDeColor(ws, row, 10, 14)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.number_format = FORMATO_PORCENTAJE_DOS_DECIMALES
        
        for col in range(16, 20):
            cell = ws.cell(row=row, column=col)
            if row == 7:
                cell = ws.cell(row=7, column=col) 
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            elif row in [8,10,16,28,32]:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.fill = SECTION_FILL
            
            if row in range(8, 33):
                if (row == 9 or (11 <= row <= 15) or (17 <= row <= 27) or (29 <= row <= 31)):
                    formatoCondicional_EstalaDeColor(ws, row, 16, 19)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.number_format = FORMATO_PORCENTAJE_DOS_DECIMALES

def FormatoAnalisis3(ws):
    ws.column_dimensions[get_column_letter(9)].width = 3
    ws.column_dimensions[get_column_letter(15)].width = 3
    ws.merge_cells('J6:N6')
    aplicarBorde(ws, 'J6:N76')
    ws['J6'].value = "Análisis Vertical"
    ws['J6'].fill = ENCABEZADO_NARANJA
    ws['J6'].font = negrita
    ws['J6'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('P6:S6')
    aplicarBorde(ws, 'P6:S76')
    ws['P6'].value = "Análisis Horizontal"
    ws['P6'].fill = ENCABEZADO_NARANJA
    ws['P6'].font = negrita
    ws['P6'].alignment = Alignment(horizontal='center', vertical='center')

    copiar_celdas(ws,ws,'D7:H7',7,10)
    copiar_celdas(ws,ws,'D7:G7',7,16)

    for row in range(7, 77):
        for col in range(10, 15):
            cell = ws.cell(row=row, column=col)
            if row == 7:
                cell = ws.cell(row=7, column=col) 
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            elif row in [8,9,15,28,29,30,43,56,57,58,65,76]:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.fill = SECTION_FILL
            
            if row in range(8, 77):
                if ((10 <= row <= 14) or (16 <= row <= 27) or (31 <= row <= 42) or (44 <= row <= 55) or (59 <= row <= 64) or (66 <= row <= 75)):
                    formatoCondicional_EstalaDeColor(ws, row, 10, 14)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.number_format = FORMATO_PORCENTAJE_DOS_DECIMALES
        
        for col in range(16, 20):
            cell = ws.cell(row=row, column=col)
            if row == 7:
                cell = ws.cell(row=7, column=col) 
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            elif row in [8,9,15,28,29,30,43,56,57,58,65,76]:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.fill = SECTION_FILL
            
            if row in range(8, 77):
                if ((10 <= row <= 14) or (16 <= row <= 27) or (31 <= row <= 42) or (44 <= row <= 55) or (59 <= row <= 64) or (66 <= row <= 75)):
                    formatoCondicional_EstalaDeColor(ws, row, 16, 19)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.number_format = FORMATO_PORCENTAJE_DOS_DECIMALES

def limpiar_rango_Libre(ws, rango_excel):
    BORDE_POR_DEFECTO = Border(left=Side(style=None), 
                               right=Side(style=None), 
                               top=Side(style=None), 
                               bottom=Side(style=None))
    try:
        min_col, min_row, max_col, max_row = range_boundaries(rango_excel)
    except Exception:
        print(f"Error: El rango '{rango_excel}' no es un formato de rango válido.")
        return

    # 3. Iterar y limpiar cada celda
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
            cell.border = BORDE_POR_DEFECTO
            cell.number_format = 'General'
            
def aplicarBorde(ws, rango_excel):
    min_col, min_row, max_col, max_row = range_boundaries(rango_excel)

    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER

def formatoCondicional_EstalaDeColor(ws, numero_fila, col_inicio, col_fin):
    # 1. Construir el string del rango de la fila
    col_inicio_letra = get_column_letter(col_inicio)
    col_fin_letra = get_column_letter(col_fin)

    #Definición de rango
    rango_fila = f"{col_inicio_letra}{numero_fila}:{col_fin_letra}{numero_fila}"
    
    # 2. Definir la regla de Escala de Color (Verde es bueno)
    regla_escala_3_colores = ColorScaleRule(
        start_color=Color('F8696B'),  # Verde (Alto)
        mid_color=Color('FFEB84'),    # Amarillo (Medio)
        end_color=Color('63BE7B'),    # Rojo (Bajo)

        start_type='percent', 
        mid_type='percent', 
        end_type='percent',
                
        start_value=100,    # Verde
        mid_value=50,       # Amarillo
        end_value=0         # Rojo
    )
    ws.conditional_formatting.add(rango_fila, regla_escala_3_colores)

def analisis_Ratios(path_xlsx):
    """Prepara estructura y calcula ratios - compatible con nombres antiguos y nuevos"""
    wb = load_workbook(path_xlsx)
    
    # Buscar hoja de ratios por nombre nuevo o antiguo
    try:
        ws = wb['Ratios Financieros']
        print("✓ Usando hoja: Ratios Financieros")
    except KeyError:
        try:
            ws = wb['Hoja5']
            print("✓ Usando hoja: Hoja5")
        except KeyError:
            print("Error: No se pudo encontrar la hoja de Ratios")
            return

    # Preparar estructura
    ws.column_dimensions[get_column_letter(9)].width = 3
    ws.column_dimensions[get_column_letter(10)].width = 30
    copiar_celdas(ws,ws,'D7:H7',7,11)
    copiar_celdas(ws,ws,'D7:H7',11,11)
    copiar_celdas(ws,ws,'D7:H7',16,11)
    copiar_celdas(ws,ws,'D7:H7',20,11)

    ws['J7'].value = "RATIOS DE LIQUIDEZ"
    ws['J7'].fill = ENCABEZADO_NARANJA
    ws['J7'].font = Font(size=11, bold=True)
    ws['J8'].value = "Liquidez Corriente"
    ws['J9'].value = "Prueba Ácida"

    ws['J11'].value = "RATIOS DE GESTIÓN"
    ws['J11'].fill = ENCABEZADO_NARANJA
    ws['J11'].font = Font(size=11, bold=True)
    ws['J12'].value = "Rotación de Cuentas por cobrar"
    ws['J13'].value = "Rotación de Inventarios"
    ws['J14'].value = "Rotación de Activos Totales"

    ws['J16'].value = "RATIOS DE ENDEUDAMIENTO"
    ws['J16'].fill = ENCABEZADO_NARANJA
    ws['J16'].font = Font(size=11, bold=True)
    ws['J17'].value = "Razón de deuda total"
    ws['J18'].value = "Razón de deuda/patrimonio"

    ws['J20'].value = "RATIOS DE RENTABILIDAD"
    ws['J20'].fill = ENCABEZADO_NARANJA
    ws['J20'].font = Font(size=11, bold=True)
    ws['J21'].value = "Margen neto"
    ws['J22'].value = "ROA"
    ws['J23'].value = "ROE"

    aplicarBorde(ws, 'J7:O9')
    aplicarBorde(ws, 'J11:O14')
    aplicarBorde(ws, 'J16:O18')
    aplicarBorde(ws, 'J20:O23')

    for row in range(6, 24):
        for col in range(11, 16):
            cell = ws.cell(row=row, column=col)
            if row in [7,11,16,20]:
                cell = ws.cell(row=row, column=col) 
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.number_format = FORMATO_NUMERICO
            else:
                cell.number_format = FORMATO_NUMERICO_FINANCIERO

    wb.save(path_xlsx)
    
    # Calcular los ratios después de preparar la estructura
    calcular_ratios_financieros_compatibles(path_xlsx)

def calcular_ratios_financieros_compatibles(path_xlsx):
    """VERSIÓN DEFINITIVA CORREGIDA"""
    wb = load_workbook(path_xlsx)
    
    # Buscar hojas necesarias
    hojas = {}
    
    for nombre in ['Estado de Situación Financiera', 'Hoja1']:
        if nombre in wb.sheetnames:
            hojas['situacion'] = wb[nombre]
            break
    
    for nombre in ['Estado de Resultados', 'Hoja2']:
        if nombre in wb.sheetnames:
            hojas['resultados'] = wb[nombre]
            break
    
    for nombre in ['Ratios Financieros', 'Hoja5']:
        if nombre in wb.sheetnames:
            hojas['ratios'] = wb[nombre]
            break
    
    if not all(k in hojas for k in ['situacion', 'resultados', 'ratios']):
        print("Error: No se encontraron todas las hojas necesarias")
        return
    
    print("✓ Todas las hojas necesarias encontradas")
    
    # DETECTAR SI ES BANCO
    es_banco = detectar_si_es_banco(hojas['situacion'], hojas['resultados'])
    tipo_empresa = "BANCO" if es_banco else "EMPRESA REGULAR"
    print(f"✓ Tipo de empresa detectado: {tipo_empresa}")
    
    # Mapeo de años a columnas
    columnas_anios = {'2024': 'D', '2023': 'E', '2022': 'F', '2021': 'G', '2020': 'H'}
    columnas_destino = {'2024': 'K', '2023': 'L', '2022': 'M', '2021': 'N', '2020': 'O'}
    
    for anio, col_origen in columnas_anios.items():
        col_destino = columnas_destino[anio]
        print(f"\n=== CALCULANDO RATIOS PARA {anio} ===")
        
        try:
            # BUSCAR VALORES CORRECTOS CON ESCALA ADECUADA
            print("Buscando valores con escala correcta...")
            
            # ACTIVO TOTAL (en millones)
            activo_total = buscar_valor_exacto(hojas['situacion'], col_origen, "TOTAL DEL ACTIVO", 51)
            if activo_total:
                print(f"✓ Activo Total: {activo_total:,.0f} millones")
            
            # PASIVO TOTAL (en millones)  
            pasivo_total = buscar_valor_exacto(hojas['situacion'], col_origen, "TOTAL DEL PASIVO", 83)
            if pasivo_total:
                print(f"✓ Pasivo Total: {pasivo_total:,.0f} millones")
            
            # PATRIMONIO TOTAL (calcular si no se encuentra)
            patrimonio_total = buscar_valor_exacto(hojas['situacion'], col_origen, "PATRIMONIO")
            if not patrimonio_total and activo_total and pasivo_total:
                patrimonio_total = activo_total - pasivo_total
                print(f"✓ Patrimonio Total (calculado): {patrimonio_total:,.0f} millones")
            
            # VENTAS NETAS = TOTAL INGRESOS POR INTERESES (en millones)
            ventas_netas = buscar_valor_exacto(hojas['resultados'], col_origen, "TOTAL INGRESOS POR INTERESES", 18)
            if ventas_netas:
                print(f"✓ Ventas Netas (Ingresos Intereses): {ventas_netas:,.0f} millones")
            
            # UTILIDAD NETA (en millones)
            utilidad_neta = buscar_valor_exacto(hojas['resultados'], col_origen, "RESULTADO NETO DEL EJERCICIO")
            if not utilidad_neta:
                utilidad_neta = buscar_valor_exacto(hojas['resultados'], col_origen, "RESULTADO DEL EJERCICIO ANTES DE IMPUESTO A LA RENTA", 77)
            if utilidad_neta:
                print(f"✓ Utilidad Neta: {utilidad_neta:,.0f} millones")
            
            # CUENTAS POR COBRAR = CARTERA (aproximar con activo total para bancos)
            cuentas_por_cobrar = buscar_valor_exacto(hojas['situacion'], col_origen, "CARTERA DE CRÉDITOS VIGENTES", 26)
            if not cuentas_por_cobrar:
                cuentas_por_cobrar = activo_total * 0.8  # Aproximación: 80% del activo es cartera
                print(f"✓ Cuentas por Cobrar (estimado 80% activo): {cuentas_por_cobrar:,.0f} millones")
            
            # GASTOS POR INTERESES (en millones)
            gastos_intereses = buscar_valor_exacto(hojas['resultados'], col_origen, "GASTOS POR INTERESES")
            if not gastos_intereses:
                # Buscar en "Adeudos y obligaciones financieras" que parece ser gastos financieros
                gastos_intereses = abs(buscar_valor_exacto(hojas['resultados'], col_origen, "ADEUDOS Y OBLIGACIONES FINANCIERAS", 23))
            if gastos_intereses:
                print(f"✓ Gastos por Intereses: {gastos_intereses:,.0f} millones")
            
            print(f"\n--- VALORES FINALES {anio} ---")
            print(f"  Activo Total: {activo_total:,.0f}M")
            print(f"  Pasivo Total: {pasivo_total:,.0f}M") 
            print(f"  Patrimonio: {patrimonio_total:,.0f}M")
            print(f"  Ventas Netas: {ventas_netas:,.0f}M")
            print(f"  Utilidad Neta: {utilidad_neta:,.0f}M")
            print(f"  Cuentas por Cobrar: {cuentas_por_cobrar:,.0f}M")
            if gastos_intereses:
                print(f"  Gastos Intereses: {gastos_intereses:,.0f}M")
            
            # CALCULAR RATIOS
            print(f"\n--- RATIOS CALCULADOS {anio} ---")
            
            # 1. LIQUIDEZ CORRIENTE (Activo/Pasivo para bancos)
            if activo_total and pasivo_total and pasivo_total != 0:
                liquidez = activo_total / pasivo_total
                hojas['ratios'][f'{col_destino}8'].value = liquidez
                print(f"✓ Liquidez Corriente: {liquidez:.2f}")
            
            # 2. PRUEBA ÁCIDA (igual para bancos)
            hojas['ratios'][f'{col_destino}9'].value = hojas['ratios'][f'{col_destino}8'].value
            print(f"✓ Prueba Ácida: {hojas['ratios'][f'{col_destino}8'].value:.2f}")
            
            # 3. ROTACIÓN DE CUENTAS POR COBRAR
            if ventas_netas and cuentas_por_cobrar and cuentas_por_cobrar != 0:
                rotacion_cartera = ventas_netas / cuentas_por_cobrar
                hojas['ratios'][f'{col_destino}12'].value = rotacion_cartera
                print(f"✓ Rotación Cartera: {rotacion_cartera:.3f}")
            
            # 4. ROTACIÓN DE INVENTARIOS (no aplicable)
            hojas['ratios'][f'{col_destino}13'].value = None
            print(f"✗ Rotación Inventarios: No aplicable")
            
            # 5. ROTACIÓN DE ACTIVOS TOTALES
            if ventas_netas and activo_total and activo_total != 0:
                rotacion_activos = ventas_netas / activo_total
                hojas['ratios'][f'{col_destino}14'].value = rotacion_activos
                print(f"✓ Rotación Activos: {rotacion_activos:.3f}")
            
            # 6. RAZÓN DE DEUDA TOTAL
            if pasivo_total and activo_total and activo_total != 0:
                razon_deuda = pasivo_total / activo_total
                hojas['ratios'][f'{col_destino}17'].value = razon_deuda
                print(f"✓ Razón Deuda Total: {razon_deuda:.1%}")
            
            # 7. RAZÓN DE DEUDA/PATRIMONIO
            if pasivo_total and patrimonio_total and patrimonio_total != 0:
                deuda_patrimonio = pasivo_total / patrimonio_total
                hojas['ratios'][f'{col_destino}18'].value = deuda_patrimonio
                print(f"✓ Razón Deuda/Patrimonio: {deuda_patrimonio:.2f}")
            
            # 8. MARGEN NETO
            if utilidad_neta and ventas_netas and ventas_netas != 0:
                margen_neto = utilidad_neta / ventas_netas
                hojas['ratios'][f'{col_destino}21'].value = margen_neto
                print(f"✓ Margen Neto: {margen_neto:.1%}")
            
            # 9. ROA
            if utilidad_neta and activo_total and activo_total != 0:
                roa = utilidad_neta / activo_total
                hojas['ratios'][f'{col_destino}22'].value = roa
                print(f"✓ ROA: {roa:.1%}")
            
            # 10. ROE
            if utilidad_neta and patrimonio_total and patrimonio_total != 0:
                roe = utilidad_neta / patrimonio_total
                hojas['ratios'][f'{col_destino}23'].value = roe
                print(f"✓ ROE: {roe:.1%}")
                
        except Exception as e:
            print(f"❌ Error en {anio}: {e}")
            continue
    
    aplicar_formatos_ratios(hojas['ratios'])
    wb.save(path_xlsx)
    print(f"\n🎉 CÁLCULO DE RATIOS COMPLETADO CORRECTAMENTE")

def buscar_valor_exacto(ws, columna, concepto, fila_esperada=None):
    """Busca valores exactos con ubicación específica"""
    # Si se conoce la fila exacta, buscar allí primero
    if fila_esperada:
        cell_value = ws[f'C{fila_esperada}'].value
        if cell_value and concepto.upper() in str(cell_value).upper():
            valor = to_number(ws[f'{columna}{fila_esperada}'].value)
            if valor not in (None, 0):
                return valor
    
    # Búsqueda en todo el rango
    for row in range(1, 120):
        cell_value = ws[f'C{row}'].value
        if cell_value and concepto.upper() in str(cell_value).upper():
            valor = to_number(ws[f'{columna}{row}'].value)
            if valor not in (None, 0):
                return valor
    
    return None

def detectar_si_es_banco(hoja_situacion, hoja_resultados):
    """Detecta si los estados financieros corresponden a un banco"""
    conceptos_bancarios = [
        "CARTERA DE CRÉDITOS", "DEPÓSITOS", "OBLIGACIONES CON EL PÚBLICO",
        "MARGEN FINANCIERO", "INTERMEDIACIÓN FINANCIERA", "PRÉSTAMOS",
        "ENCAJES", "RESERVAS TÉCNICAS", "CARTERA VIGENTE"
    ]
    
    # Buscar en el balance
    for row in range(1, 100):
        cell_value = hoja_situacion[f'C{row}'].value
        if cell_value:
            cell_upper = str(cell_value).upper()
            for concepto in conceptos_bancarios:
                if concepto in cell_upper:
                    print(f"✓ Detectado concepto bancario: '{cell_value}'")
                    return True
    
    # Buscar en estado de resultados
    for row in range(1, 100):
        cell_value = hoja_resultados[f'C{row}'].value
        if cell_value:
            cell_upper = str(cell_value).upper()
            for concepto in conceptos_bancarios:
                if concepto in cell_upper:
                    print(f"✓ Detectado concepto bancario: '{cell_value}'")
                    return True
    
    print("✓ Empresa detectada como EMPRESA REGULAR (no banco)")
    return False

def calcular_ratios_empresa_regular(hojas, col_origen, col_destino, anio):
    """Calcula ratios para empresas regulares (con inventarios)"""
    print("Buscando valores para empresa regular...")
    
    # ACTIVO CORRIENTE
    activo_corriente = encontrar_valor_empresa(hojas['situacion'], col_origen, [
        "TOTAL ACTIVO CORRIENTE", "ACTIVO CORRIENTE", "TOTAL DEL ACTIVO CORRIENTE",
        "ACTIVOS CORRIENTES"
    ], "Activo Corriente")
    
    # PASIVO CORRIENTE
    pasivo_corriente = encontrar_valor_empresa(hojas['situacion'], col_origen, [
        "TOTAL PASIVO CORRIENTE", "PASIVO CORRIENTE", "TOTAL DEL PASIVO CORRIENTE",
        "PASIVOS CORRIENTES"
    ], "Pasivo Corriente")
    
    # INVENTARIOS
    inventarios = encontrar_valor_empresa(hojas['situacion'], col_origen, [
        "INVENTARIOS", "EXISTENCIAS", "INVENTARIO", "MERCADERÍAS",
        "MATERIAS PRIMAS", "PRODUCTOS EN PROCESO", "PRODUCTOS TERMINADOS"
    ], "Inventarios")
    
    # VENTAS NETAS
    ventas_netas = encontrar_valor_empresa(hojas['resultados'], col_origen, [
        "VENTAS NETAS", "INGRESOS POR VENTAS", "TOTAL VENTAS", "VENTAS",
        "INGRESOS DE ACTIVIDADES ORDINARIAS", "INGRESOS OPERACIONALES"
    ], "Ventas Netas")
    
    # CUENTAS POR COBRAR
    cuentas_por_cobrar = encontrar_valor_empresa(hojas['situacion'], col_origen, [
        "CUENTAS POR COBRAR", "CUENTAS POR COBRAR COMERCIALES", "CLIENTES",
        "DEUDORES COMERCIALES"
    ], "Cuentas por Cobrar")
    
    # COSTO DE VENTAS
    costo_ventas = encontrar_valor_empresa(hojas['resultados'], col_origen, [
        "COSTO DE VENTAS", "COSTO DE MERCADERÍA VENDIDA", "COSTO DE LO VENDIDO",
        "COSTO DE VENTA", "COSTO DE LOS INGRESOS"
    ], "Costo de Ventas")
    
    # ACTIVO TOTAL
    activo_total = encontrar_valor_empresa(hojas['situacion'], col_origen, [
        "TOTAL DEL ACTIVO", "TOTAL ACTIVO", "ACTIVO TOTAL"
    ], "Activo Total")
    
    # PASIVO TOTAL
    pasivo_total = encontrar_valor_empresa(hojas['situacion'], col_origen, [
        "TOTAL DEL PASIVO", "TOTAL PASIVO", "PASIVO TOTAL"
    ], "Pasivo Total")
    
    # PATRIMONIO TOTAL
    patrimonio_total = encontrar_valor_empresa(hojas['situacion'], col_origen, [
        "TOTAL DEL PATRIMONIO", "TOTAL PATRIMONIO", "PATRIMONIO TOTAL",
        "PATRIMONIO", "CAPITAL CONTABLE"
    ], "Patrimonio Total")
    
    # Si no encuentra patrimonio, calcularlo
    if not patrimonio_total and activo_total and pasivo_total:
        patrimonio_total = activo_total - pasivo_total
        print(f"  ✓ Patrimonio Total (calculado): {patrimonio_total:,.0f}")
    
    # UTILIDAD NETA
    utilidad_neta = encontrar_valor_empresa(hojas['resultados'], col_origen, [
        "UTILIDAD NETA", "RESULTADO DEL EJERCICIO", "RESULTADO NETO",
        "UTILIDAD DEL EJERCICIO", "BENEFICIO NETO"
    ], "Utilidad Neta")
    
    print(f"\n--- CALCULANDO RATIOS EMPRESA REGULAR PARA {anio} ---")
    
    # 1. LIQUIDEZ CORRIENTE
    if activo_corriente and pasivo_corriente and pasivo_corriente != 0:
        liquidez_corriente = activo_corriente / pasivo_corriente
        hojas['ratios'][f'{col_destino}8'].value = liquidez_corriente
        print(f"✓ Liquidez Corriente: {liquidez_corriente:.2f}")
    else:
        print(f"✗ Liquidez Corriente: No calculable")
        hojas['ratios'][f'{col_destino}8'].value = None
    
    # 2. PRUEBA ÁCIDA
    if activo_corriente and inventarios and pasivo_corriente and pasivo_corriente != 0:
        prueba_acida = (activo_corriente - inventarios) / pasivo_corriente
        hojas['ratios'][f'{col_destino}9'].value = prueba_acida
        print(f"✓ Prueba Ácida: {prueba_acida:.2f}")
    else:
        print(f"✗ Prueba Ácida: No calculable")
        hojas['ratios'][f'{col_destino}9'].value = None
    
    # 3. ROTACIÓN DE CUENTAS POR COBRAR
    if ventas_netas and cuentas_por_cobrar and cuentas_por_cobrar != 0:
        rotacion_cuentas_cobrar = ventas_netas / cuentas_por_cobrar
        hojas['ratios'][f'{col_destino}12'].value = rotacion_cuentas_cobrar
        print(f"✓ Rotación Cuentas por Cobrar: {rotacion_cuentas_cobrar:.2f}")
    else:
        print(f"✗ Rotación Cuentas por Cobrar: No calculable")
        hojas['ratios'][f'{col_destino}12'].value = None
    
    # 4. ROTACIÓN DE INVENTARIOS
    if costo_ventas and inventarios and inventarios != 0:
        rotacion_inventarios = costo_ventas / inventarios
        hojas['ratios'][f'{col_destino}13'].value = rotacion_inventarios
        print(f"✓ Rotación Inventarios: {rotacion_inventarios:.2f}")
    else:
        print(f"✗ Rotación Inventarios: No calculable")
        hojas['ratios'][f'{col_destino}13'].value = None
    
    # 5. ROTACIÓN DE ACTIVOS TOTALES
    if ventas_netas and activo_total and activo_total != 0:
        rotacion_activos = ventas_netas / activo_total
        hojas['ratios'][f'{col_destino}14'].value = rotacion_activos
        print(f"✓ Rotación Activos Totales: {rotacion_activos:.2f}")
    else:
        print(f"✗ Rotación Activos Totales: No calculable")
        hojas['ratios'][f'{col_destino}14'].value = None
    
    # 6. RAZÓN DE DEUDA TOTAL
    if pasivo_total and activo_total and activo_total != 0:
        razon_deuda_total = pasivo_total / activo_total
        hojas['ratios'][f'{col_destino}17'].value = razon_deuda_total
        print(f"✓ Razón Deuda Total: {razon_deuda_total:.2%}")
    else:
        print(f"✗ Razón Deuda Total: No calculable")
        hojas['ratios'][f'{col_destino}17'].value = None
    
    # 7. RAZÓN DE DEUDA/PATRIMONIO
    if pasivo_total and patrimonio_total and patrimonio_total != 0:
        razon_deuda_patrimonio = pasivo_total / patrimonio_total
        hojas['ratios'][f'{col_destino}18'].value = razon_deuda_patrimonio
        print(f"✓ Razón Deuda/Patrimonio: {razon_deuda_patrimonio:.2f}")
    else:
        print(f"✗ Razón Deuda/Patrimonio: No calculable")
        hojas['ratios'][f'{col_destino}18'].value = None
    
    # 8. MARGEN NETO
    if utilidad_neta and ventas_netas and ventas_netas != 0:
        margen_neto = utilidad_neta / ventas_netas
        hojas['ratios'][f'{col_destino}21'].value = margen_neto
        print(f"✓ Margen Neto: {margen_neto:.2%}")
    else:
        print(f"✗ Margen Neto: No calculable")
        hojas['ratios'][f'{col_destino}21'].value = None
    
    # 9. ROA
    if utilidad_neta and activo_total and activo_total != 0:
        roa = utilidad_neta / activo_total
        hojas['ratios'][f'{col_destino}22'].value = roa
        print(f"✓ ROA: {roa:.2%}")
    else:
        print(f"✗ ROA: No calculable")
        hojas['ratios'][f'{col_destino}22'].value = None
    
    # 10. ROE
    if utilidad_neta and patrimonio_total and patrimonio_total != 0:
        roe = utilidad_neta / patrimonio_total
        hojas['ratios'][f'{col_destino}23'].value = roe
        print(f"✓ ROE: {roe:.2%}")
    else:
        print(f"✗ ROE: No calculable")
        hojas['ratios'][f'{col_destino}23'].value = None

def calcular_ratios_banco(hojas, col_origen, col_destino, anio):
    """Calcula ratios específicos para bancos - VERSIÓN CORREGIDA"""
    print("Buscando valores para banco...")
    
    # ACTIVO CORRIENTE - Para bancos usar Activo Total como aproximación
    activo_corriente = encontrar_valor_empresa(hojas['situacion'], col_origen, [
        "TOTAL DEL ACTIVO", "TOTAL ACTIVO"
    ], "Activo Corriente (Total Activo)")
    
    # PASIVO CORRIENTE - Para bancos usar Pasivo Total como aproximación  
    pasivo_corriente = encontrar_valor_empresa(hojas['situacion'], col_origen, [
        "TOTAL DEL PASIVO", "TOTAL PASIVO"
    ], "Pasivo Corriente (Total Pasivo)")
    
    # VENTAS NETAS = TOTAL INGRESOS POR INTERESES (CORREGIDO)
    ventas_netas = encontrar_valor_empresa(hojas['resultados'], col_origen, [
        "TOTAL INGRESOS POR INTERESES", "INGRESOS POR INTERESES", 
        "INTERESES GANADOS", "INGRESOS FINANCIEROS"
    ], "Ventas Netas (Ingresos Intereses)")
    
    # Si no encuentra, buscar en las filas donde sabemos que existe
    if not ventas_netas:
        # Buscar específicamente en la fila C18 donde sabemos está el total
        for row in [18, 19, 20, 17, 16]:
            cell_value = hojas['resultados'][f'C{row}'].value
            if cell_value and "TOTAL INGRESOS POR INTERESES" in str(cell_value).upper():
                ventas_netas = to_number(hojas['resultados'][f'{col_origen}{row}'].value)
                if ventas_netas:
                    print(f"  ✓ Ventas Netas (forzado C{row}): {ventas_netas:,.0f}")
                    break
    
    # CUENTAS POR COBRAR = CARTERA DE CRÉDITOS
    cuentas_por_cobrar = encontrar_valor_empresa(hojas['situacion'], col_origen, [
        "CARTERA DE CRÉDITOS VIGENTES", "CARTERA VIGENTE", "CRÉDITOS VIGENTES",
        "PRÉSTAMOS", "CARTERA DE CRÉDITO"
    ], "Cuentas por Cobrar (Cartera)")
    
    # Si no encuentra cartera específica, usar activo total como aproximación
    if not cuentas_por_cobrar:
        cuentas_por_cobrar = activo_corriente
        print(f"  ✓ Cuentas por Cobrar (usando Activo Total): {cuentas_por_cobrar:,.0f}")
    
    # COSTO DE VENTAS = GASTOS POR INTERESES (CORREGIDO)
    costo_ventas = encontrar_valor_empresa(hojas['resultados'], col_origen, [
        "GASTOS POR INTERESES", "INTERESES PAGADOS", "GASTOS FINANCIEROS"
    ], "Costo de Ventas (Gastos Intereses)")
    
    # ACTIVO TOTAL
    activo_total = activo_corriente  # Ya lo tenemos
    
    # PASIVO TOTAL  
    pasivo_total = pasivo_corriente  # Ya lo tenemos
    
    # PATRIMONIO TOTAL
    patrimonio_total = encontrar_valor_empresa(hojas['situacion'], col_origen, [
        "TOTAL DEL PATRIMONIO", "TOTAL PATRIMONIO", "PATRIMONIO"
    ], "Patrimonio Total")
    
    # Si no encuentra patrimonio, calcularlo correctamente
    if not patrimonio_total and activo_total and pasivo_total:
        patrimonio_total = activo_total - pasivo_total
        print(f"  ✓ Patrimonio Total (calculado): {patrimonio_total:,.0f}")
    
    # UTILIDAD NETA
    utilidad_neta = encontrar_valor_empresa(hojas['resultados'], col_origen, [
        "UTILIDAD NETA", "RESULTADO DEL EJERCICIO", "RESULTADO NETO",
        "RESULTADO DEL EJERCICIO ANTES DE IMPUESTO A LA RENTA"
    ], "Utilidad Neta")
    
    print(f"\n--- VALORES CORREGIDOS PARA {anio} ---")
    print(f"  Activo Total: {activo_total:,.0f}")
    print(f"  Pasivo Total: {pasivo_total:,.0f}") 
    print(f"  Patrimonio Total: {patrimonio_total:,.0f}")
    print(f"  Ventas Netas (Ingresos Intereses): {ventas_netas:,.0f}" if ventas_netas else "  Ventas Netas: No encontrado")
    print(f"  Utilidad Neta: {utilidad_neta:,.0f}")
    
    print(f"\n--- CALCULANDO RATIOS BANCO CORREGIDOS PARA {anio} ---")
    
    # 1. LIQUIDEZ CORRIENTE (Activo Total / Pasivo Total para bancos)
    if activo_total and pasivo_total and pasivo_total != 0:
        liquidez_corriente = activo_total / pasivo_total
        hojas['ratios'][f'{col_destino}8'].value = liquidez_corriente
        print(f"✓ Liquidez Corriente: {liquidez_corriente:.2f}")
    else:
        print(f"✗ Liquidez Corriente: No calculable")
        hojas['ratios'][f'{col_destino}8'].value = None
    
    # 2. PRUEBA ÁCIDA (igual a liquidez para bancos)
    hojas['ratios'][f'{col_destino}9'].value = hojas['ratios'][f'{col_destino}8'].value
    print(f"✓ Prueba Ácida: {hojas['ratios'][f'{col_destino}8'].value:.2f}")
    
    # 3. ROTACIÓN DE CUENTAS POR COBRAR
    if ventas_netas and cuentas_por_cobrar and cuentas_por_cobrar != 0:
        rotacion_cuentas_cobrar = ventas_netas / cuentas_por_cobrar
        hojas['ratios'][f'{col_destino}12'].value = rotacion_cuentas_cobrar
        print(f"✓ Rotación Cuentas por Cobrar: {rotacion_cuentas_cobrar:.4f}")
    else:
        print(f"✗ Rotación Cuentas por Cobrar: No calculable")
        hojas['ratios'][f'{col_destino}12'].value = None
    
    # 4. ROTACIÓN DE INVENTARIOS (no aplicable)
    hojas['ratios'][f'{col_destino}13'].value = None
    print(f"✗ Rotación Inventarios: No aplicable para banco")
    
    # 5. ROTACIÓN DE ACTIVOS TOTALES
    if ventas_netas and activo_total and activo_total != 0:
        rotacion_activos = ventas_netas / activo_total
        hojas['ratios'][f'{col_destino}14'].value = rotacion_activos
        print(f"✓ Rotación Activos Totales: {rotacion_activos:.4f}")
    else:
        print(f"✗ Rotación Activos Totales: No calculable")
        hojas['ratios'][f'{col_destino}14'].value = None
    
    # 6. RAZÓN DE DEUDA TOTAL
    if pasivo_total and activo_total and activo_total != 0:
        razon_deuda_total = pasivo_total / activo_total
        hojas['ratios'][f'{col_destino}17'].value = razon_deuda_total
        print(f"✓ Razón Deuda Total: {razon_deuda_total:.2%}")
    else:
        print(f"✗ Razón Deuda Total: No calculable")
        hojas['ratios'][f'{col_destino}17'].value = None
    
    # 7. RAZÓN DE DEUDA/PATRIMONIO
    if pasivo_total and patrimonio_total and patrimonio_total != 0:
        razon_deuda_patrimonio = pasivo_total / patrimonio_total
        hojas['ratios'][f'{col_destino}18'].value = razon_deuda_patrimonio
        print(f"✓ Razón Deuda/Patrimonio: {razon_deuda_patrimonio:.2f}")
    elif pasivo_total and patrimonio_total and patrimonio_total < 0:
        # Si patrimonio es negativo, el ratio también será negativo
        razon_deuda_patrimonio = pasivo_total / patrimonio_total
        hojas['ratios'][f'{col_destino}18'].value = razon_deuda_patrimonio
        print(f"✓ Razón Deuda/Patrimonio: {razon_deuda_patrimonio:.2f} (patrimonio negativo)")
    else:
        print(f"✗ Razón Deuda/Patrimonio: No calculable")
        hojas['ratios'][f'{col_destino}18'].value = None
    
    # 8. MARGEN NETO
    if utilidad_neta and ventas_netas and ventas_netas != 0:
        margen_neto = utilidad_neta / ventas_netas
        hojas['ratios'][f'{col_destino}21'].value = margen_neto
        print(f"✓ Margen Neto: {margen_neto:.2%}")
    else:
        print(f"✗ Margen Neto: No calculable")
        hojas['ratios'][f'{col_destino}21'].value = None
    
    # 9. ROA
    if utilidad_neta and activo_total and activo_total != 0:
        roa = utilidad_neta / activo_total
        hojas['ratios'][f'{col_destino}22'].value = roa
        print(f"✓ ROA: {roa:.2%}")
    else:
        print(f"✗ ROA: No calculable")
        hojas['ratios'][f'{col_destino}22'].value = None
    
    # 10. ROE
    if utilidad_neta and patrimonio_total and patrimonio_total != 0:
        roe = utilidad_neta / patrimonio_total
        hojas['ratios'][f'{col_destino}23'].value = roe
        print(f"✓ ROE: {roe:.2%}")
    elif utilidad_neta and patrimonio_total and patrimonio_total < 0:
        # Si patrimonio es negativo, ROE también será negativo
        roe = utilidad_neta / patrimonio_total
        hojas['ratios'][f'{col_destino}23'].value = roe
        print(f"✓ ROE: {roe:.2%} (patrimonio negativo)")
    else:
        print(f"✗ ROE: No calculable")
        hojas['ratios'][f'{col_destino}23'].value = None

def calcular_ratios_comunes(hojas, col_destino, activo_corriente, pasivo_corriente, inventarios,
                           ventas_netas, cuentas_por_cobrar, costo_ventas, activo_total,
                           pasivo_total, patrimonio_total, utilidad_neta, es_banco=False):
    """Calcula los ratios comunes para ambos tipos de empresa"""
    
    # 1. LIQUIDEZ CORRIENTE
    if activo_corriente and pasivo_corriente and pasivo_corriente != 0:
        liquidez_corriente = activo_corriente / pasivo_corriente
        hojas['ratios'][f'{col_destino}8'].value = liquidez_corriente
        tipo = "Banco" if es_banco else "Empresa"
        print(f"✓ Liquidez Corriente ({tipo}): {liquidez_corriente:.2f}")
    else:
        print(f"✗ Liquidez Corriente: No calculable")
        hojas['ratios'][f'{col_destino}8'].value = None
    
    # 2. PRUEBA ÁCIDA
    if es_banco:
        # Para bancos, prueba ácida = liquidez corriente (no hay inventarios)
        if activo_corriente and pasivo_corriente and pasivo_corriente != 0:
            prueba_acida = activo_corriente / pasivo_corriente
            hojas['ratios'][f'{col_destino}9'].value = prueba_acida
            print(f"✓ Prueba Ácida (Banco): {prueba_acida:.2f}")
        else:
            hojas['ratios'][f'{col_destino}9'].value = None
    else:
        # Para empresas regulares, fórmula normal
        if activo_corriente and inventarios and pasivo_corriente and pasivo_corriente != 0:
            prueba_acida = (activo_corriente - inventarios) / pasivo_corriente
            hojas['ratios'][f'{col_destino}9'].value = prueba_acida
            print(f"✓ Prueba Ácida (Empresa): {prueba_acida:.2f}")
        else:
            hojas['ratios'][f'{col_destino}9'].value = None
    
    # 3. ROTACIÓN DE CUENTAS POR COBRAR
    if ventas_netas and cuentas_por_cobrar and cuentas_por_cobrar != 0:
        rotacion_cuentas_cobrar = ventas_netas / cuentas_por_cobrar
        hojas['ratios'][f'{col_destino}12'].value = rotacion_cuentas_cobrar
        print(f"✓ Rotación Cuentas por Cobrar: {rotacion_cuentas_cobrar:.2f}")
    else:
        hojas['ratios'][f'{col_destino}12'].value = None
    
    # 4. ROTACIÓN DE INVENTARIOS (solo para empresas regulares)
    if not es_banco:
        if costo_ventas and inventarios and inventarios != 0:
            rotacion_inventarios = costo_ventas / inventarios
            hojas['ratios'][f'{col_destino}13'].value = rotacion_inventarios
            print(f"✓ Rotación Inventarios: {rotacion_inventarios:.2f}")
        else:
            hojas['ratios'][f'{col_destino}13'].value = None
    else:
        hojas['ratios'][f'{col_destino}13'].value = None
        print(f"✗ Rotación Inventarios: No aplicable para banco")

    # 5. ROTACIÓN DE ACTIVOS TOTALES
    if ventas_netas and activo_total and activo_total != 0:
        rotacion_activos = ventas_netas / activo_total
        hojas['ratios'][f'{col_destino}14'].value = rotacion_activos
        print(f"✓ Rotación Activos Totales: {rotacion_activos:.2f}")
    else:
        print(f"✗ Rotación Activos Totales: No calculable")
        hojas['ratios'][f'{col_destino}14'].value = None
    
    # 6. RAZÓN DE DEUDA TOTAL
    if pasivo_total and activo_total and activo_total != 0:
        razon_deuda_total = pasivo_total / activo_total
        hojas['ratios'][f'{col_destino}17'].value = razon_deuda_total
        print(f"✓ Razón Deuda Total: {razon_deuda_total:.2%}")
    else:
        print(f"✗ Razón Deuda Total: No calculable")
        hojas['ratios'][f'{col_destino}17'].value = None
    
    # 7. RAZÓN DE DEUDA/PATRIMONIO
    if pasivo_total and patrimonio_total and patrimonio_total != 0:
        razon_deuda_patrimonio = pasivo_total / patrimonio_total
        hojas['ratios'][f'{col_destino}18'].value = razon_deuda_patrimonio
        print(f"✓ Razón Deuda/Patrimonio: {razon_deuda_patrimonio:.2f}")
    else:
        print(f"✗ Razón Deuda/Patrimonio: No calculable")
        hojas['ratios'][f'{col_destino}18'].value = None
    
    # 8. MARGEN NETO
    if utilidad_neta and ventas_netas and ventas_netas != 0:
        margen_neto = utilidad_neta / ventas_netas
        hojas['ratios'][f'{col_destino}21'].value = margen_neto
        print(f"✓ Margen Neto: {margen_neto:.2%}")
    else:
        print(f"✗ Margen Neto: No calculable")
        hojas['ratios'][f'{col_destino}21'].value = None
    
    # 9. ROA
    if utilidad_neta and activo_total and activo_total != 0:
        roa = utilidad_neta / activo_total
        hojas['ratios'][f'{col_destino}22'].value = roa
        print(f"✓ ROA: {roa:.2%}")
    else:
        print(f"✗ ROA: No calculable")
        hojas['ratios'][f'{col_destino}22'].value = None
    
    # 10. ROE
    if utilidad_neta and patrimonio_total and patrimonio_total != 0:
        roe = utilidad_neta / patrimonio_total
        hojas['ratios'][f'{col_destino}23'].value = roe
        print(f"✓ ROE: {roe:.2%}")
    else:
        print(f"✗ ROE: No calculable")
        hojas['ratios'][f'{col_destino}23'].value = None

    # Los ratios 5-10 son iguales para ambos...
    # (Resto de ratios se mantienen igual)

def encontrar_valor_empresa(ws, columna, conceptos, nombre_concepto=""):
    """Función mejorada para buscar valores"""
    for row in range(1, 150):
        cell_value = ws[f'C{row}'].value
        if cell_value:
            cell_upper = str(cell_value).upper().strip()
            for concepto in conceptos:
                if concepto.upper() in cell_upper:
                    valor = to_number(ws[f'{columna}{row}'].value)
                    if valor not in (None, 0):
                        print(f"  ✓ {nombre_concepto}: {valor:,.0f} (C{row}: '{cell_value}')")
                        return valor
    
    print(f"  ✗ {nombre_concepto}: No encontrado")
    return None

def encontrar_valor_banco(ws, columna, conceptos, nombre_concepto=""):
    """Función específica para buscar valores bancarios"""
    for row in range(1, 150):
        cell_value = ws[f'C{row}'].value
        if cell_value:
            cell_upper = str(cell_value).upper().strip()
            for concepto in conceptos:
                if concepto.upper() in cell_upper:
                    valor = to_number(ws[f'{columna}{row}'].value)
                    if valor not in (None, 0):
                        print(f"  ✓ {nombre_concepto}: {valor:,.0f} (C{row}: '{cell_value}')")
                        return valor
    
    print(f"  ✗ {nombre_concepto}: No encontrado")
    return None

def encontrar_valor_mejorado(ws, columna, conceptos, nombre_concepto=""):
    """Función mejorada para buscar valores con debugging"""
    for row in range(1, 150):  # Buscar en un rango más amplio
        cell_value = ws[f'C{row}'].value
        if cell_value:
            cell_upper = str(cell_value).upper().strip()
            for concepto in conceptos:
                if concepto.upper() in cell_upper:
                    valor = to_number(ws[f'{columna}{row}'].value)
                    if valor not in (None, 0):
                        print(f"  ✓ {nombre_concepto}: {valor:,.0f} (C{row}: '{cell_value}')")
                        return valor
    
    # Búsqueda alternativa más flexible
    for row in range(1, 150):
        cell_value = ws[f'C{row}'].value
        if cell_value:
            cell_upper = str(cell_value).upper().strip()
            # Buscar por palabras clave del primer concepto
            if conceptos:
                palabras_clave = conceptos[0].upper().split()
                coincidencias = sum(1 for palabra in palabras_clave if palabra in cell_upper)
                if coincidencias >= len(palabras_clave) - 1:  # Al menos n-1 palabras coinciden
                    valor = to_number(ws[f'{columna}{row}'].value)
                    if valor not in (None, 0):
                        print(f"  ✓ {nombre_concepto} (flexible): {valor:,.0f} (C{row}: '{cell_value}')")
                        return valor
    
    print(f"  ✗ {nombre_concepto}: No encontrado")
    return None

def aplicar_formatos_ratios(ws):
    """Aplica formatos específicos a los ratios financieros"""
    
    # Formato para ratios de liquidez y gestión (2 decimales)
    for row in range(8, 15):  # Liquidez y Gestión
        for col in ['K', 'L', 'M', 'N', 'O']:
            cell = ws[f'{col}{row}']
            if cell.value is not None:
                cell.number_format = "0.00"
    
    # Formato para ratios de endeudamiento (porcentaje)
    for row in range(17, 19):  # Endeudamiento
        for col in ['K', 'L', 'M', 'N', 'O']:
            cell = ws[f'{col}{row}']
            if cell.value is not None:
                cell.number_format = "0.00%"
    
    # Formato para ratios de rentabilidad (porcentaje)
    for row in range(21, 24):  # Rentabilidad
        for col in ['K', 'L', 'M', 'N', 'O']:
            cell = ws[f'{col}{row}']
            if cell.value is not None:
                cell.number_format = "0.00%"







def graficosRatios(path_xlsx):
    wb = load_workbook(path_xlsx)
    
    # Crear hoja para gráficos
    if 'Hoja6' in wb.sheetnames:
        GraRati = wb['Hoja6']
    else:
        GraRati = wb.create_sheet(title="Hoja6", index=None)

    GrafRati = wb['Hoja5']

    ws = GraRati

    dir_path = os.path.dirname(path_xlsx)
    nombre = os.path.basename(dir_path)
    nombreEmpresa = nombre.replace('_', ' ')

    # Configuración inicial de la hoja
    ws['A1'].value = "GRÁFICOS DE RATIOS"
    ws['A1'].font = fuente_titulo
    ws['A3'].value = "Periodo: Anual"
    ws['A3'].font = negrita
    ws['A4'].value = f"Empresa: {nombreEmpresa}"
    ws['A4'].font = negrita
    ws['A5'].value = "Tipo: Individual"
    ws['A5'].font = negrita

    # Configurar dimensiones de columnas
    ws.column_dimensions[get_column_letter(1)].width = 3
    ws.column_dimensions[get_column_letter(2)].width = 3
    for i in range(3, 30):
        ws.column_dimensions[get_column_letter(i)].width = 15
    for i in [5, 8, 11, 14]:
        ws.column_dimensions[get_column_letter(i)].width = 32
    print("Hola bro")
    # Crear datos de ejemplo para los ratios
    datos_ratios = extraer_ratios(GrafRati, fila_inicio=7, fila_fin=23, col_inicio=11, col_fin=15)
    print(datos_ratios)
    rangos = {
    "Liquidez Corriente": ("C10", "D10"),  
    "Prueba Ácida": ("F10", "G10"),
    "Rotación de Cuentas por cobrar": ("I10", "J10"),
    "Rotación de Inventarios": ("L10", "M10"),
    "Rotación de Activos Totales": ("O10", "P10"),
    "Razón de deuda total": ("C33", "D33"),
    "Razón de deuda/patrimonio": ("F33", "G33"),
    "Margen neto": ("I33", "J33"),
    "ROA": ("L33", "M33"),
    "ROE": ("O33", "P33")
    }

    anios = ["2024", "2023", "2022", "2021", "2020"]

    for indicador, (col_anio, col_valor) in rangos.items():
        col_letra_anio, fila_inicio = coordinate_from_string(col_anio)
        col_letra_valor, _ = coordinate_from_string(col_valor)

        for i, anio in enumerate(anios):
            fila = fila_inicio + i
            # Insertar año
            ws[f"{col_letra_anio}{fila}"].value = anio
            # Insertar valor (si existe en el diccionario)
            valor = datos_ratios.get(indicador, {}).get(anio, None)
            ws[f"{col_letra_valor}{fila}"].value = valor


    df = pd.DataFrame(datos_ratios).T  # Transponer para que los indicadores sean filas
    df.index.name = "Indicador"
    df = df.reset_index()
    
    # Pivotear para que quede Año como columna
    df = df.melt(id_vars=["Indicador"], var_name="Año", value_name="Valor")
    
    # Crear columnas individuales para cada indicador
    df_pivot = df.pivot(index="Año", columns="Indicador", values="Valor").reset_index()
    
    # Renombrar columnas para que coincidan con las que usas en gráficos
    df_pivot.rename(columns={
        "Liquidez Corriente": "Liquidez_Corriente",
        "Prueba Ácida": "Prueba_Acida",
        "Rotación de Cuentas por cobrar": "Rotacion_CtasCobrar",
        "Rotación de Inventarios": "Rotacion_Inventarios",
        "Rotación de Activos Totales": "Rotacion_ActivosTotales",
        "Razón de deuda total": "Razon_DeudaTotal",
        "Razón de deuda/patrimonio": "Razon_DeudaPatrimonio",
        "Margen neto": "Margen_Neto",
        "ROA": "ROA",
        "ROE": "ROE"
    }, inplace=True)

    try:
        # Crear gráficos en memoria
        crear_grafico_liquidez(df_pivot, ws)
        crear_grafico_gestion(df_pivot, ws)
        crear_grafico_endeudamiento(df_pivot, ws)
        crear_grafico_rentabilidad(df_pivot, ws)

        # Aplicar formatos y bordes
        aplicar_formatos_tablas(ws)
        
    except Exception as e:
        print(f"Error al crear gráficos: {e}")
        # En caso de error, al menos guardar la estructura básica
        aplicar_formatos_tablas(ws)
        ws['C16'].value = f"Error al generar gráficos: {str(e)}"
        ws['C16'].font = Font(color="FF0000", italic=True)
    
    wb.save(path_xlsx)

def crear_grafico_liquidez(df, ws):
    """Crear gráfico para ratios de liquidez - Versión corregida"""
    try:
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 5))
        print("gaaaaaaa")
        # Gráfico de Liquidez Corriente
        ax1.bar(df['Año'], df['Liquidez_Corriente'], color=['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7'])
        ax1.set_title('Liquidez Corriente', fontweight='bold')
        ax1.set_ylabel('Ratio')
        ax1.grid(True, alpha=0.3)
        
        # Gráfico de Prueba Ácida
        ax2.bar(df['Año'], df['Prueba_Acida'], color=['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7'])
        ax2.set_title('Prueba Ácida', fontweight='bold')
        ax2.set_ylabel('Ratio')
        ax2.grid(True, alpha=0.3)
        
        plt.tight_layout()
        
        # SOLUCIÓN: Guardar en archivo temporal y NO cerrarlo inmediatamente
        temp_dir = tempfile.gettempdir()
        img_path = os.path.join(temp_dir, f'grafico_liquidez_{os.getpid()}.png')
        plt.savefig(img_path, format='png', dpi=150, bbox_inches='tight')
        plt.close(fig)
        
        # Insertar imagen desde archivo
        img = Image(img_path)
        img.width = 600
        img.height = 300
        ws.add_image(img, 'C15')
        
        # El archivo se mantendrá hasta que Excel lo procese
        
    except Exception as e:
        print(f"Error en gráfico de liquidez: {e}")
        plt.close('all')

def crear_grafico_gestion(df, ws):
    """Crear gráfico para ratios de gestión - Versión corregida"""
    try:
        fig, (ax1, ax2, ax3) = plt.subplots(1, 3, figsize=(15, 5))
        
        # Gráfico de Rotación de Cuentas por Cobrar
        ax1.plot(df['Año'], df['Rotacion_CtasCobrar'], marker='o', linewidth=2, color='#FF6B6B')
        ax1.set_title('Rotación Cuentas por Cobrar', fontweight='bold')
        ax1.set_ylabel('Veces')
        ax1.grid(True, alpha=0.3)
        
        # Gráfico de Rotación de Inventarios
        ax2.plot(df['Año'], df['Rotacion_Inventarios'], marker='s', linewidth=2, color='#4ECDC4')
        ax2.set_title('Rotación de Inventarios', fontweight='bold')
        ax2.set_ylabel('Veces')
        ax2.grid(True, alpha=0.3)
        
        # Gráfico de Rotación de Activos Totales
        ax3.plot(df['Año'], df['Rotacion_ActivosTotales'], marker='^', linewidth=2, color='#45B7D1')
        ax3.set_title('Rotación Activos Totales', fontweight='bold')
        ax3.set_ylabel('Veces')
        ax3.grid(True, alpha=0.3)
        
        plt.tight_layout()
        
        temp_dir = tempfile.gettempdir()
        img_path = os.path.join(temp_dir, f'grafico_gestion_{os.getpid()}.png')
        plt.savefig(img_path, format='png', dpi=150, bbox_inches='tight')
        plt.close(fig)
        
        img = Image(img_path)
        img.width = 800
        img.height = 300
        ws.add_image(img, 'I15')
        
    except Exception as e:
        print(f"Error en gráfico de gestión: {e}")
        plt.close('all')

def crear_grafico_endeudamiento(df, ws):
    """Crear gráfico para ratios de endeudamiento - Versión corregida"""
    try:
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 5))
        
        # Gráfico de Razón de Deuda Total
        ax1.bar(df['Año'], df['Razon_DeudaTotal'], color=['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7'])
        ax1.set_title('Razón de Deuda Total', fontweight='bold')
        ax1.set_ylabel('Ratio')
        ax1.set_ylim(0, 1)
        ax1.grid(True, alpha=0.3)
        
        # Gráfico de Razón Deuda/Patrimonio
        ax2.bar(df['Año'], df['Razon_DeudaPatrimonio'], color=['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7'])
        ax2.set_title('Razón Deuda/Patrimonio', fontweight='bold')
        ax2.set_ylabel('Ratio')
        ax2.grid(True, alpha=0.3)
        
        plt.tight_layout()
        
        temp_dir = tempfile.gettempdir()
        img_path = os.path.join(temp_dir, f'grafico_endeudamiento_{os.getpid()}.png')
        plt.savefig(img_path, format='png', dpi=150, bbox_inches='tight')
        plt.close(fig)
        
        img = Image(img_path)
        img.width = 600
        img.height = 300
        ws.add_image(img, 'C39')
        
    except Exception as e:
        print(f"Error en gráfico de endeudamiento: {e}")
        plt.close('all')

def crear_grafico_rentabilidad(df, ws):
    """Crear gráfico para ratios de rentabilidad con estilo profesional"""
    try:
        # Estilo general
        plt.style.use('seaborn-v0_8-whitegrid')  # limpio y moderno

        # Crear figura con tres subgráficos
        fig, (ax1, ax2, ax3) = plt.subplots(1, 3, figsize=(16, 6))
        fig.suptitle("Ratios de Rentabilidad", fontsize=16, fontweight="bold", color="#2C3E50")

        # Convertir a porcentaje
        margen_neto_pct = df['Margen_Neto'] * 100
        roa_pct = df['ROA'] * 100
        roe_pct = df['ROE'] * 100

        # Paleta corporativa (azules + gris neutro)
        colores = ['#2980B9', '#3498DB', '#5DADE2', '#A9CCE3', '#D5DBDB']

        # -------- Margen Neto --------
        ax1.bar(df['Año'], margen_neto_pct, color=colores, edgecolor="black", linewidth=0.8)
        ax1.set_title("Margen Neto", fontsize=13, fontweight="bold", color="#2C3E50")
        ax1.set_ylabel("Porcentaje (%)", fontsize=11)
        ax1.tick_params(axis="x", rotation=45)

        # -------- ROA --------
        ax2.bar(df['Año'], roa_pct, color=colores, edgecolor="black", linewidth=0.8)
        ax2.set_title("ROA", fontsize=13, fontweight="bold", color="#2C3E50")
        ax2.set_ylabel("Porcentaje (%)", fontsize=11)
        ax2.tick_params(axis="x", rotation=45)

        # -------- ROE --------
        ax3.bar(df['Año'], roe_pct, color=colores, edgecolor="black", linewidth=0.8)
        ax3.set_title("ROE", fontsize=13, fontweight="bold", color="#2C3E50")
        ax3.set_ylabel("Porcentaje (%)", fontsize=11)
        ax3.tick_params(axis="x", rotation=45)

        # Ajustar espaciado
        plt.tight_layout(rect=[0, 0, 1, 0.95])

        # Guardar imagen temporalmente
        temp_dir = tempfile.gettempdir()
        img_path = os.path.join(temp_dir, f'grafico_rentabilidad_{os.getpid()}.png')
        plt.savefig(img_path, format="png", dpi=180, bbox_inches="tight")
        plt.close(fig)

        # Insertar en Excel
        img = Image(img_path)
        img.width = 800   # más ancho
        img.height = 350  # más alto
        ws.add_image(img, "I39")

    except Exception as e:
        print(f"Error en gráfico de rentabilidad: {e}")
        plt.close("all")

def aplicar_formatos_tablas(ws):
    """Aplicar formatos a las tablas de ratios"""
    # Ratios de Liquidez
    ws.merge_cells('C7:G7')
    ws['C7'].value = "RATIOS DE LIQUIDEZ"
    ws['C7'].fill = ENCABEZADO_PURPURA
    ws['C7'].font = Font(size=12, bold=True, color="FFFFFF")
    
    # Configurar tablas
    configurar_tabla_liquidez(ws)
    configurar_tabla_gestion(ws)
    configurar_tabla_endeudamiento(ws)
    configurar_tabla_rentabilidad(ws)
    
    # Aplicar bordes
    aplicarBorde(ws, 'C7:G14')
    centrar_rango(ws, 'C7:G14')
    aplicarBorde(ws, 'I7:P13')
    centrar_rango(ws, 'I7:P13')
    aplicarBorde(ws, 'C30:G37')
    centrar_rango(ws, 'C30:G37')
    aplicarBorde(ws, 'I30:P37')
    centrar_rango(ws, 'I30:P37')

def configurar_tabla_rentabilidad(ws):
    ws.merge_cells('I30:P30')
    ws.merge_cells('K31:K37')
    ws.merge_cells('N31:N37')
    ws['I30'].value = "RATIOS DE RENTABILIDAD"
    ws['I30'].fill = ENCABEZADO_PURPURA
    ws['I30'].font = HEADER_FONT
    ws['I30'].font = Font(size=12, bold=True)
    ws.merge_cells('I31:J31')
    ws['I31'].value = "Margen neto"
    ws['I31'].fill = ENCABEZADO_NARANJA
    ws['I32'].value = "Año"
    ws['I32'].fill = ENCABEZADO_CELESTE
    ws['J32'].value = "Valor"
    ws['J32'].fill = ENCABEZADO_CELESTE
    ws.merge_cells('L31:M31')
    ws['L31'].value = "ROA"
    ws['L31'].fill = ENCABEZADO_NARANJA
    ws['L32'].value = "Año"
    ws['L32'].fill = ENCABEZADO_CELESTE
    ws['M32'].value = "Valor"
    ws['M32'].fill = ENCABEZADO_CELESTE
    ws.merge_cells('O31:P31')
    ws['O31'].value = "ROE"
    ws['O31'].fill = ENCABEZADO_NARANJA
    ws['O32'].value = "Año"
    ws['O32'].fill = ENCABEZADO_CELESTE
    ws['P32'].value = "Valor"
    ws['P32'].fill = ENCABEZADO_CELESTE

def configurar_tabla_liquidez(ws):
    ws.merge_cells('I7:P7')
    ws.merge_cells('K8:K13')
    ws.merge_cells('N8:N13')
    ws['I7'].value = "RATIOS DE GESTIÓN"
    ws['I7'].fill = ENCABEZADO_PURPURA
    ws['I7'].font = HEADER_FONT
    ws['I7'].font = Font(size=12, bold=True)
    ws.merge_cells('I8:J8')
    ws['I8'].value = "Rotación de Cuentas por cobrar"
    ws['I8'].fill = ENCABEZADO_NARANJA
    ws['I9'].value = "Año"
    ws['I9'].fill =ENCABEZADO_CELESTE
    ws['J9'].value = "Valor"
    ws['J9'].fill =ENCABEZADO_CELESTE
    ws.merge_cells('L8:M8')
    ws['L8'].value = "Rotación de Inventarios"
    ws['L8'].fill = ENCABEZADO_NARANJA
    ws['L9'].value = "Año"
    ws['L9'].fill =ENCABEZADO_CELESTE
    ws['M9'].value = "Valor"
    ws['M9'].fill =ENCABEZADO_CELESTE
    ws.merge_cells('O8:P8')
    ws['O8'].value = "Rotación de Activos Totales"
    ws['O8'].fill = ENCABEZADO_NARANJA
    ws['O9'].value = "Año"
    ws['O9'].fill = ENCABEZADO_CELESTE
    ws['P9'].value = "Valor"
    ws['P9'].fill = ENCABEZADO_CELESTE

def configurar_tabla_gestion(ws):
    """Configurar tabla de ratios de liquidez"""
    # Mantener tu estructura original de tablas
    ws.merge_cells('C8:D8')
    ws['C8'].value = "Liquidez Corriente"
    ws['C8'].fill = ENCABEZADO_NARANJA
    ws['C9'].value = "Año"
    ws['C9'].fill = ENCABEZADO_CELESTE
    ws['D9'].value = "Valor"
    ws['D9'].fill = ENCABEZADO_CELESTE
    
    ws.merge_cells('F8:G8')
    ws['F8'].value = "Prueba Ácida"
    ws['F8'].fill = ENCABEZADO_NARANJA
    ws['F9'].value = "Año"
    ws['F9'].fill = ENCABEZADO_CELESTE
    ws['G9'].value = "Valor"
    ws['G9'].fill = ENCABEZADO_CELESTE

def configurar_tabla_endeudamiento(ws):
    ws.merge_cells('C30:G30')
    ws.merge_cells('E31:E37')
    ws['C30'].value = "RATIOS DE ENDEUDAMIENTO"
    ws['C30'].fill = ENCABEZADO_PURPURA
    ws['C30'].font = HEADER_FONT
    ws['C30'].font = Font(size=12, bold=True)
    ws.merge_cells('C31:D31')
    ws['C31'].value = "Razón de deuda total"
    ws['C31'].fill = ENCABEZADO_NARANJA
    ws['C32'].value = "Año"
    ws['C32'].fill = ENCABEZADO_CELESTE
    ws['D32'].value = "Valor"
    ws['D32'].fill = ENCABEZADO_CELESTE
    ws.merge_cells('F31:G31')
    ws['F31'].value = "Razón de deuda/patrimonio"
    ws['F31'].fill = ENCABEZADO_NARANJA
    ws['F32'].value = "Año"
    ws['F32'].fill = ENCABEZADO_CELESTE
    ws['G32'].value = "Valor"
    ws['G32'].fill = ENCABEZADO_CELESTE



def centrar_rango(ws, rango):
    """Centrar contenido en un rango de celdas"""
    min_col, min_row, max_col, max_row = range_boundaries(rango)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')



def renombrar(path_xlsx):
    wb = load_workbook(path_xlsx)
    sistFinan = wb['Hoja1']
    resultados = wb['Hoja2']
    patrimonio = wb['Hoja3']
    flujos = wb['Hoja4']
    ratios = wb['Hoja5']
    graratios = wb['Hoja6']

    sistFinan.title = "Estado de Situación Financiera"
    resultados.title = "Estado de Resultados"
    patrimonio.title = "Estado de Patrimonio Neto"
    flujos.title = "Estado de Flujo de Efectivo"
    ratios.title = "Ratios Financieros"
    graratios.title = "Gráficos de Ratios Financieros"

    wb.save(path_xlsx)


def centrar_rango(ws, rango):
    alineacion_centrada = Alignment(horizontal='center', vertical='center')
    min_col, min_row, max_col, max_row = range_boundaries(rango)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = alineacion_centrada